from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import tempfile
from datetime import datetime
from decimal import Decimal, ROUND_DOWN
from typing import List, Dict, Any
from pathlib import Path
from copy import copy
from src.config.folder_mappings import settings
from src.utils.chinese_pdf import format_case_number

# ── 台灣縣市鄉鎮郵遞區號靜態映射（3 位數字）──────────────────────────────────────
_TAIWAN_POSTAL_CODES: Dict[str, Dict[str, str]] = {
    "臺北市": {
        "中正區": "100", "大同區": "103", "中山區": "104", "松山區": "105",
        "大安區": "106", "萬華區": "108", "信義區": "110", "士林區": "111",
        "北投區": "112", "內湖區": "114", "南港區": "115", "文山區": "116",
    },
    "基隆市": {
        "仁愛區": "200", "信義區": "201", "中正區": "202", "中山區": "203",
        "安樂區": "204", "暖暖區": "205", "七堵區": "206",
    },
    "新北市": {
        "萬里區": "207", "金山區": "208", "板橋區": "220", "汐止區": "221",
        "深坑區": "222", "石碇區": "223", "瑞芳區": "224", "平溪區": "226",
        "雙溪區": "227", "貢寮區": "228", "新店區": "231", "坪林區": "232",
        "烏來區": "233", "永和區": "234", "中和區": "235", "土城區": "236",
        "三峽區": "237", "樹林區": "238", "鶯歌區": "239", "三重區": "241",
        "新莊區": "242", "泰山區": "243", "林口區": "244", "蘆洲區": "247",
        "五股區": "248", "八里區": "249", "淡水區": "251", "三芝區": "252",
        "石門區": "253",
    },
    "宜蘭縣": {
        "宜蘭市": "260", "頭城鎮": "261", "礁溪鄉": "262", "壯圍鄉": "263",
        "員山鄉": "264", "羅東鎮": "265", "三星鄉": "266", "大同鄉": "267",
        "五結鄉": "268", "冬山鄉": "269", "蘇澳鎮": "270", "南澳鄉": "272",
    },
    "新竹市": {
        "東區": "300", "北區": "300", "香山區": "300",
    },
    "新竹縣": {
        "竹北市": "302", "湖口鄉": "303", "新豐鄉": "304", "新埔鎮": "305",
        "關西鎮": "306", "芎林鄉": "307", "寶山鄉": "308", "竹東鎮": "310",
        "五峰鄉": "311", "橫山鄉": "312", "尖石鄉": "313", "北埔鄉": "314",
        "峨眉鄉": "315",
    },
    "桃園市": {
        "中壢區": "320", "平鎮區": "324", "龍潭區": "325", "楊梅區": "326",
        "新屋區": "327", "觀音區": "328", "桃園區": "330", "龜山區": "333",
        "八德區": "334", "大溪區": "335", "復興區": "336", "大園區": "337",
        "蘆竹區": "338",
    },
    "苗栗縣": {
        "竹南鎮": "350", "頭份市": "351", "三灣鄉": "352", "南庄鄉": "353",
        "獅潭鄉": "354", "後龍鎮": "356", "通霄鎮": "357", "苑裡鎮": "358",
        "苗栗市": "360", "造橋鄉": "361", "頭屋鄉": "362", "公館鄉": "363",
        "大湖鄉": "364", "泰安鄉": "365", "銅鑼鄉": "366", "三義鄉": "367",
        "西湖鄉": "368", "卓蘭鎮": "369",
    },
    "臺中市": {
        "中區": "400", "東區": "401", "南區": "402", "西區": "403",
        "北區": "404", "北屯區": "406", "西屯區": "407", "南屯區": "408",
        "太平區": "411", "大里區": "412", "霧峰區": "413", "烏日區": "414",
        "豐原區": "420", "后里區": "421", "石岡區": "422", "東勢區": "423",
        "和平區": "424", "新社區": "426", "潭子區": "427", "大雅區": "428",
        "神岡區": "429", "大肚區": "432", "沙鹿區": "433", "龍井區": "434",
        "梧棲區": "435", "清水區": "436", "大甲區": "437", "外埔區": "438",
        "大安區": "439",
    },
    "彰化縣": {
        "彰化市": "500", "芬園鄉": "502", "花壇鄉": "503", "秀水鄉": "504",
        "鹿港鎮": "505", "福興鄉": "506", "線西鄉": "507", "和美鎮": "508",
        "伸港鄉": "509", "員林市": "510", "社頭鄉": "511", "永靖鄉": "512",
        "埔心鄉": "513", "溪湖鎮": "514", "大村鄉": "515", "埔鹽鄉": "516",
        "田中鎮": "520", "北斗鎮": "521", "田尾鄉": "522", "埤頭鄉": "523",
        "溪州鄉": "524", "竹塘鄉": "525", "二林鎮": "526", "大城鄉": "527",
        "芳苑鄉": "528", "二水鄉": "530",
    },
    "南投縣": {
        "南投市": "540", "中寮鄉": "541", "草屯鎮": "542", "國姓鄉": "544",
        "埔里鎮": "545", "仁愛鄉": "546", "名間鄉": "551", "集集鎮": "552",
        "水里鄉": "553", "魚池鄉": "555", "信義鄉": "556", "竹山鎮": "557",
        "鹿谷鄉": "558",
    },
    "嘉義市": {
        "東區": "600", "西區": "600",
    },
    "嘉義縣": {
        "番路鄉": "602", "梅山鄉": "603", "竹崎鄉": "604", "阿里山鄉": "605",
        "中埔鄉": "606", "大埔鄉": "607", "水上鄉": "608", "鹿草鄉": "611",
        "太保市": "612", "朴子市": "613", "東石鄉": "614", "六腳鄉": "615",
        "新港鄉": "616", "民雄鄉": "621", "大林鎮": "622", "溪口鄉": "623",
        "義竹鄉": "624", "布袋鎮": "625",
    },
    "雲林縣": {
        "斗南鎮": "630", "大埤鄉": "631", "虎尾鎮": "632", "土庫鎮": "633",
        "褒忠鄉": "634", "東勢鄉": "635", "台西鄉": "636", "崙背鄉": "637",
        "麥寮鄉": "638", "斗六市": "640", "林內鄉": "643", "古坑鄉": "646",
        "莿桐鄉": "647", "西螺鎮": "648", "二崙鄉": "649", "北港鎮": "651",
        "水林鄉": "652", "口湖鄉": "653", "四湖鄉": "654", "元長鄉": "655",
    },
    "臺南市": {
        "中西區": "700", "東區": "701", "南區": "702", "北區": "704",
        "安平區": "708", "安南區": "709", "永康區": "710", "歸仁區": "711",
        "新化區": "712", "左鎮區": "713", "玉井區": "714", "楠西區": "715",
        "南化區": "716", "仁德區": "717", "關廟區": "718", "龍崎區": "719",
        "官田區": "720", "麻豆區": "721", "佳里區": "722", "西港區": "723",
        "七股區": "724", "將軍區": "725", "學甲區": "726", "北門區": "727",
        "新營區": "730", "後壁區": "731", "白河區": "732", "東山區": "733",
        "六甲區": "734", "下營區": "735", "柳營區": "736", "鹽水區": "737",
        "善化區": "741", "大內區": "742", "山上區": "743", "新市區": "744",
        "安定區": "745",
    },
    "高雄市": {
        "新興區": "800", "前金區": "801", "苓雅區": "802", "鹽埕區": "803",
        "鼓山區": "804", "旗津區": "805", "前鎮區": "806", "三民區": "807",
        "楠梓區": "811", "小港區": "812", "左營區": "813", "仁武區": "814",
        "大社區": "815", "岡山區": "820", "路竹區": "821", "阿蓮區": "822",
        "田寮區": "823", "燕巢區": "824", "橋頭區": "825", "梓官區": "826",
        "彌陀區": "827", "永安區": "828", "湖內區": "829", "鳳山區": "830",
        "大寮區": "831", "林園區": "832", "鳥松區": "833", "大樹區": "840",
        "旗山區": "842", "美濃區": "843", "六龜區": "844", "內門區": "845",
        "杉林區": "846", "甲仙區": "847", "桃源區": "848", "那瑪夏區": "849",
        "茂林區": "851", "茄萣區": "852",
    },
    "屏東縣": {
        "屏東市": "900", "三地門鄉": "901", "霧台鄉": "902", "瑪家鄉": "903",
        "九如鄉": "904", "里港鄉": "905", "高樹鄉": "906", "鹽埔鄉": "907",
        "長治鄉": "908", "麟洛鄉": "909", "竹田鄉": "911", "內埔鄉": "912",
        "萬丹鄉": "913", "潮州鎮": "920", "泰武鄉": "921", "來義鄉": "922",
        "萬巒鄉": "923", "崁頂鄉": "924", "新埤鄉": "925", "南州鄉": "926",
        "林邊鄉": "927", "東港鎮": "928", "琉球鄉": "929", "佳冬鄉": "931",
        "新園鄉": "932", "枋寮鄉": "940", "枋山鄉": "941", "春日鄉": "942",
        "獅子鄉": "943", "車城鄉": "944", "牡丹鄉": "945", "恆春鎮": "946",
        "滿州鄉": "947",
    },
    "臺東縣": {
        "臺東市": "950", "綠島鄉": "951", "蘭嶼鄉": "952", "延平鄉": "953",
        "卑南鄉": "954", "鹿野鄉": "955", "關山鎮": "956", "海端鄉": "957",
        "池上鄉": "958", "東河鄉": "959", "成功鎮": "961", "長濱鄉": "962",
        "太麻里鄉": "963", "金峰鄉": "964", "大武鄉": "965", "達仁鄉": "966",
    },
    "花蓮縣": {
        "花蓮市": "970", "新城鄉": "971", "秀林鄉": "972", "吉安鄉": "973",
        "壽豐鄉": "974", "鳳林鎮": "975", "光復鄉": "976", "豐濱鄉": "977",
        "瑞穗鄉": "978", "萬榮鄉": "979", "玉里鎮": "981", "卓溪鄉": "982",
        "富里鄉": "983",
    },
    "澎湖縣": {
        "馬公市": "880", "西嶼鄉": "881", "望安鄉": "882", "七美鄉": "883",
        "白沙鄉": "884", "湖西鄉": "885",
    },
    "金門縣": {
        "金城鎮": "890", "金湖鎮": "891", "金沙鎮": "892", "金寧鄉": "893",
        "烈嶼鄉": "894", "烏坵鄉": "896",
    },
    "連江縣": {
        "南竿鄉": "209", "北竿鄉": "210", "莒光鄉": "211", "東引鄉": "212",
    },
}

# ── 共用樣式常數（全報表共用）────────────────────────────────────────────────────
_XL_S_M = Side(style='medium')
_XL_S_T = Side(style='thin')

# 字型
_XL_F16  = Font(name='微軟正黑體', size=16)
_XL_F16B = Font(name='微軟正黑體', size=16, bold=True)
_XL_F12  = Font(name='微軟正黑體', size=12)
_XL_F14  = Font(name='微軟正黑體', size=14)

# 對齊
_XL_A_CTR  = Alignment(horizontal='center', vertical='center')
_XL_A_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
_XL_A_RT   = Alignment(horizontal='right', vertical='center')
_XL_A_NOTE = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 列高
_XL_ROW_H_TITLE = 83.0
_XL_ROW_H_DATE  = 25.0
_XL_ROW_H_HDR1  = 25.0
_XL_ROW_H_HDR2  = 45.0
_XL_ROW_H_DATA  = 21.0
_XL_ROW_H_TOTAL = 21.0
_XL_ROW_H_NOTE  = 96.0

# ── A04/A05 規則建構常數（去範本化）──────────────────────────────────────────────
# 備註文字（A04/A05 共用）
_A0405_NOTE_TEXT = (
    '備註 :\n'
    '1、受理「補助案件」如跨越縣市/鄉鎮區土地，以第一筆縣市/鄉鎮區作為補助案件數之統計基準。\n'
)

# 每個年度欄頭的三格邊框（_tpl_merge_horizontal 使用；先設框再合併保留右外框）
_A0405_YEAR_HDR_BORDERS = [
    Border(left=_XL_S_M, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_T),  # 首欄
    Border(left=None,     right=None,     top=_XL_S_M, bottom=_XL_S_T),   # 中欄
    Border(left=None,     right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_T),  # 末欄
]

# 子欄頭文字與邊框（補助案件數 / 補助面積 / 補助金額）
_A0405_SUBHDR_TEXT = ['補助案件數\n(已結案)', '補助面積(公頃)', '補助金額(元)']
_A0405_SUBHDR_BORDER = [
    Border(left=_XL_S_M, right=_XL_S_T, bottom=_XL_S_M),
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_M),
    Border(left=_XL_S_T, right=_XL_S_M, bottom=_XL_S_M),
]

# 資料格邊框（年度子欄，thin 分隔線）
_A0405_DATA_YEAR_BORDER = [
    Border(left=_XL_S_M, right=_XL_S_T, bottom=_XL_S_T),  # 案件數
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 面積
    Border(left=_XL_S_T, right=_XL_S_M, bottom=_XL_S_T),  # 金額
]

# 固定欄頭邊框（key = (fixed_cols, col_0based)）
_A0405_FIXED_HDR_BORDER = {
    (2, 0): Border(left=_XL_S_M, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # A04 縣市
    (2, 1): Border(left=_XL_S_T, right=None,     top=_XL_S_M, bottom=_XL_S_M),  # A04 鄉鎮區
    (1, 0): Border(left=_XL_S_M, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),  # A05 管理處
}

# 固定欄 HDR2 底端格邊框（垂直合併後 _tpl_force_interior_border 注入）
_A0405_FIXED_HDR2_BORDER = {
    (2, 0): Border(left=_XL_S_M, right=_XL_S_T, bottom=_XL_S_M),
    (2, 1): Border(left=_XL_S_T, right=None,     bottom=_XL_S_M),
    (1, 0): Border(left=_XL_S_M, right=_XL_S_M, bottom=_XL_S_M),
}

# 固定欄資料格邊框
_A0405_FIXED_DATA_BORDER = {
    (2, 0): Border(left=_XL_S_M, right=_XL_S_T, bottom=_XL_S_T),
    (2, 1): Border(left=_XL_S_T, right=None,     bottom=_XL_S_T),
    (1, 0): Border(left=_XL_S_M, right=None,     bottom=_XL_S_T),
}

# 欄寬（fixed_cols → {col_0based: width, 'year': width_per_year_col}）
_A0405_COL_WIDTHS = {
    2: {0: 24.33, 1: 24.33, 'year': 24.33},  # A04 (fixed_cols=2)
    1: {0: 35.33,            'year': 24.33},  # A05 (fixed_cols=1)
}

# ── A02/A03/A07 規則建構常數（去範本化）─────────────────────────────────────────
# 表頭邊框（col_count, col_0based）
_A0203_HDR_BORDER = {
    # A02/A07 (5欄): 縣市 | 鄉鎮區 | 補助案件數 | 補助面積 | 補助金額
    (5, 0): Border(left=_XL_S_M, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),
    (5, 1): Border(left=_XL_S_T, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),
    (5, 2): Border(left=None,     right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),
    (5, 3): Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),
    (5, 4): Border(left=_XL_S_T, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),
    # A03 (4欄): 管理處 | 補助案件數 | 補助面積 | 補助金額
    (4, 0): Border(left=_XL_S_M, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),
    (4, 1): Border(left=None,     right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),
    (4, 2): Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),
    (4, 3): Border(left=_XL_S_T, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),
}
# 資料格邊框（col_count, col_0based）
_A0203_DATA_BORDER = {
    (5, 0): Border(left=_XL_S_M, right=_XL_S_T, bottom=_XL_S_T),
    (5, 1): Border(left=_XL_S_T, right=_XL_S_M, bottom=_XL_S_T),
    (5, 2): Border(left=None,     right=_XL_S_T, bottom=_XL_S_T),
    (5, 3): Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),
    (5, 4): Border(left=_XL_S_T, right=_XL_S_M, bottom=_XL_S_T),
    (4, 0): Border(left=_XL_S_M, right=_XL_S_M, bottom=_XL_S_T),
    (4, 1): Border(left=None,     right=_XL_S_T, bottom=_XL_S_T),
    (4, 2): Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),
    (4, 3): Border(left=_XL_S_T, right=_XL_S_M, bottom=_XL_S_T),
}
# 欄寬（col_count → {col_0based: width}）
_A0203_COL_WIDTHS = {
    5: {0: 24.33, 1: 24.33, 2: 24.33, 3: 24.33, 4: 24.33},  # A02/A07 (5欄)
    4: {0: 35.33, 1: 24.33, 2: 24.33, 3: 24.33},             # A03 (4欄)
}
# 表頭文字（col_count → [col 名稱]）
_A0203_HDR_TEXT = {
    5: ['縣市', '鄉鎮區', '補助案件數\n(已結案)', '補助面積(公頃)', '補助金額(元)'],
    4: ['管理處', '補助案件數\n(已結案)', '補助面積(公頃)', '補助金額(元)'],
}
# 表頭對齊（補助案件數含換行需 wrap_text）
_A0203_HDR_ALIGN = {
    5: [_XL_A_CTR, _XL_A_CTR, _XL_A_WRAP, _XL_A_CTR, _XL_A_CTR],
    4: [_XL_A_CTR, _XL_A_WRAP, _XL_A_CTR, _XL_A_CTR],
}

# ── A06 規則建構常數（去範本化）────────────────────────────────────────────────
_A06_COL_WIDTHS = {1: 35.33, **{i: 29 for i in range(2, 13)}}  # A欄35.33，其餘24.33
# 數字格式（1-based，與欄位對應）
_A06_NUM_FMT = {
    1:  'General',
    2:  'General',
    3:  '#,##0_);[Red](#,##0)',
    4:  'General',
    5:  'General',
    6:  '#,##0_);[Red](#,##0)',
    7:  '#,##0_ ;[Red]\\-#,##0\\ ',
    8:  'General',
    9:  'General',
    10: '#,##0_);[Red](#,##0)',
    11: '0.00_);[Red]\\(0.00\\)',
    12: '0.00_);[Red]\\(0.00\\)',
}
# 欄頭文字（Row 3，12 欄）
_A06_HDR_TEXT = [
    '管理處', '預定執行面積(公頃)', '預定執行預算(元)', '已編預算案件數',
    '已編預算面積(公頃)', '已編列補助款(元)', '未編列補助款(元)',
    '已驗收案件數', '已驗收面積(公頃)', '已驗收金額(元)', '面積執行率%', '計畫執行率%',
]
# 欄頭邊框（Row 3，0-based index；col 6/11 右側 None，視覺由相鄰格提供）
_A06_HDR_BORDER = [
    Border(left=_XL_S_M, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 0 管理處
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 1 預定面積
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 2 預定預算
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 3 已編案件數
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 4 已編面積
    Border(left=_XL_S_T, right=None,     top=_XL_S_M, bottom=_XL_S_M),  # 5 已編補助款
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 6 未編補助款
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 7 已驗案件數
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 8 已驗面積
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 9 已驗金額
    Border(left=_XL_S_T, right=None,     top=_XL_S_M, bottom=_XL_S_M),  # 10 面積執行率
    Border(left=_XL_S_T, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),  # 11 計畫執行率
]
# 資料格邊框（0-based index；底線 thin，合計列改用 medium）
_A06_DATA_BORDER = [
    Border(left=_XL_S_M, right=_XL_S_T, bottom=_XL_S_T),  # 0
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 1
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 2
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 3
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 4
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 5
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 6
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 7
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 8
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 9
    Border(left=_XL_S_T, right=None,     bottom=_XL_S_T),  # 10
    Border(left=_XL_S_T, right=_XL_S_M, bottom=_XL_S_T),  # 11
]
_A06_ROW_H_NOTE = 110.0
_A06_NOTE_TEXT = (
    '備註：\n'
    '    1、受理「補助案件」如跨越縣市/鄉鎮區土地，以第一筆縣市/鄉鎮區作為補助案件數之統計基準。\n'
    '計算公式 :\n'
    '    1、計畫執行率%=已編列補助款/預定執行預算。\n'
    '    2、面積執行率%=已編預算面積/預定執行面積。'
)
# Rich Text runs（underline: bool, text: str）；preserve 由 _inject_note_rich_text 自動判斷
_A06_NOTE_RUNS = (
    (False, '備註：\n    1、受理「補助案件」如跨越縣市/鄉鎮區土地，以第一筆縣市/鄉鎮區作為'),
    (True,  '補助案件數'),
    (False, '之統計基準。\n'),
    (True,  '計算公式 :'),
    (False, '\n    1、計畫執行率%=已編列補助款/預定執行預算。\n    2、面積執行率%=已編預算面積/預定執行面積。'),
)


# ── A01 規則建構常數（去範本化）────────────────────────────────────────────────
_A01_COL_WIDTHS = {1: 35.33, 2: 20.5, 3: 23.0, 4: 23.0, 5: 20.5, 6: 31.33}
_A01_HDR_TEXT = [
    '管理處', '核定金額(元)', '補助案件數\n(已結案)', '補助面積(公頃)', '補助金額(元)', '補助款\n執行率%',
]
_A01_HDR_ALIGN = [_XL_A_CTR, _XL_A_CTR, _XL_A_WRAP, _XL_A_CTR, _XL_A_CTR, _XL_A_WRAP]
_A01_HDR_BORDER = [
    Border(left=_XL_S_M, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),  # 0 管理處
    Border(left=None,     right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 1 核定金額
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 2 補助案件數
    Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_M, bottom=_XL_S_M),  # 3 補助面積
    Border(left=_XL_S_T, right=None,     top=_XL_S_M, bottom=_XL_S_M),  # 4 補助金額
    Border(left=_XL_S_T, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M),  # 5 執行率
]
_A01_DATA_BORDER = [
    Border(left=_XL_S_M, right=_XL_S_M, bottom=_XL_S_T),  # 0 管理處（獨立欄，兩側 medium）
    Border(left=None,     right=_XL_S_T, bottom=_XL_S_T),  # 1 核定金額
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 2 補助案件數
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 3 補助面積
    Border(left=_XL_S_T, right=_XL_S_T, bottom=_XL_S_T),  # 4 補助金額
    Border(left=_XL_S_T, right=_XL_S_M, bottom=_XL_S_T),  # 5 執行率
]
_A01_NUM_FMT = {1: 'General', 2: '#,##0_ ', 3: 'General', 4: 'General', 5: '#,##0_ ', 6: '0.00%'}
_A01_NOTE_TEXT = (
    '備註 :\n'
    '1、受理「補助案件」如跨越縣市/鄉鎮區土地，以第一筆縣市/鄉鎮區作為補助案件數之統計基準。\n'
    '2、計算公式：補助款執行率%=補助金額/核定金額。'
)
# Rich Text runs（underline: bool, text: str）；preserve 由 _inject_note_rich_text 自動判斷
_A01_NOTE_RUNS = (
    (False, '備註 :\n1、受理「補助案件」如跨越縣市/鄉鎮區土地，以第一筆縣市/鄉鎮區作為'),
    (True,  '補助案件數'),
    (False, '之統計基準。\n'),
    (True,  '計算公式：\n'),
    (False, '1、補助款執行率%=補助金額/核定金額。'),
)


# ── A07 規則建構常數（框線/欄寬/列高與 A02/A03 共用 _A0203_*/_XL_ROW_H_* 常數）──
_A07_NUM_FMT = {1: 'General', 2: 'General', 3: 'General', 4: 'General', 5: '#,##0'}
_A07_NOTE_TEXT = (
    '備註 :\n'
    '1、受理「補助案件」如跨越縣市/鄉鎮區土地，以第一筆縣市/鄉鎮區作為補助案件數之統計基準。'
)


# ── A09/A10 規則建構常數（事業區域內外推動成果統計，無範本依賴）────────────────

# 欄寬（當年度 16 欄 / 非當年度 7 欄，含 col 1 標籤欄）
_A0910_COL_WIDTHS = {
    16: {1: 27.0, **{i: 20.0 for i in range(2, 17)}},
    7:  {1: 27.0, **{i: 20.0 for i in range(2, 8)}},
}

# 表頭填充色
_A0910_FILL_BUDGETED  = PatternFill('solid', fgColor='E2EFDA')  # 已編列 淡綠
_A0910_FILL_COMPLETED = PatternFill('solid', fgColor='DDEBF7')  # 已結案 淡藍
_A0910_FILL_SUBTOTAL  = PatternFill('solid', fgColor='F2F2F2')  # 小計 淡灰

# 數字格式（每組 3 欄依序：案件數, 面積, 金額）
_A0910_NUM_FMT = ['General', '0.0000', '#,##0']

# 備註文字
_A0910_NOTE_TEXT = (
    '備註：\n'
    '1、事業區域歸屬依「任一土地」規則：補助案件所屬土地中任一筆 isIrrigationArea 為真，則整案歸入事業區域內，否則歸入事業區域外。\n'
    '2、「已結案」係指 status 為 completed 或 submitted 之案件；「已編列」係指案件狀態為 under review, ompleted 或 submitted 之所有有效案件。'
)
_A0910_ROW_H_NOTE = 60.0


# ── B03 規則建構常數（各縣市鄉鎮區各類補助項目統計表，19欄）────────────────────
# 欄寬（1-based，A=1 ~ S=19）
_B03_COL_WIDTHS = {
    1: 18.0,  # A 縣市
    2: 18.0,  # B 鄉鎮區
    3: 20.0,  # C 灌溉型式
    4: 14.0,  # D 補助面積(公頃)
    5: 12.0,  # E 補助案件數
    6: 16.0,  # F 農戶配合款(元)
    7: 16.0,  # G 田間管路設施(元)
    8: 16.0,  # H 調蓄設施(元)
    9: 16.0,  # I 動力設施(元)
    10: 16.0, # J 調控設備(元)
    11: 14.0, # K 設計費(元)
    12: 16.0, # L 總計(元)
    13: 16.0, # M 工程經費合計(元)
    14: 12.0, # N 噸
    15: 10.0,  # O 座
    16: 12.0, # P 抽水機(臺)
    17: 16.0, # Q 補助款(元)
    18: 12.0, # R 百分比%
    19: 16.0, # S 總工程費(元)
}
# 數值格式（1-based）
_B03_NUM_FMT = {
    1:  'General',    # 縣市
    2:  'General',    # 鄉鎮區
    3:  'General',    # 灌溉型式
    4:  '0.000',      # 補助面積(公頃)
    5:  '#,##0',      # 補助案件數
    6:  '#,##0',      # 農戶配合款(元)
    7:  '#,##0',      # 田間管路設施(元)
    8:  '#,##0',      # 調蓄設施(元)
    9:  '#,##0',      # 動力設施(元)
    10: '#,##0',      # 調控設備(元)
    11: '#,##0',      # 設計費(元)
    12: '#,##0',      # 補助總計(元)
    13: '#,##0',      # 工程經費合計(元)
    14: '#,##0',      # 噸
    15: '#,##0',      # 座
    16: '#,##0',      # 抽水機(臺)
    17: '#,##0',      # 補助款(元/公頃)
    18: '0.00%',      # 百分比%（值為 0-1 小數，openpyxl 自動 ×100 顯示）
    19: '#,##0',      # 總工程費(元/公頃)
}
# 備註文字
_B03_NOTE_TEXT = (
    '備註：\n'
    '1、受理「補助案件」如跨越縣市/鄉鎮區土地，以第一筆縣市/鄉鎮區作為補助案件數之統計基準。\n'
    '2、工程經費合計 = 農戶配合款 + 補助經費總計。\n'
    '3、*調蓄設施蓄水池欄位：噸的數量為蓄水池座數容量的總和。'
)

# ── 外出拍攝照片攜帶表（PCF）常數 ──────────────────────────────────────────────
_PCF_COL_WIDTHS = {1: 12, 2: 15, 3: 12, 4: 15, 5: 40, 6: 18, 7: 16, 8: 20, 9: 24, 10: 18, 11: 60}
_PCF_HDR_TEXT = ['案件編號', '申請人姓名', '鄉鎮', '段名', '地號', '面積（公頃）', '設施類型', '末端型式', '農作物', '電話', '通訊地址']
_PCF_DATA_ROWS_PER_PAGE = 16  # 每頁資料列數
_PCF_THIN_BORDER = Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_T, bottom=_XL_S_T)
_PCF_HDR_FONT = Font(name='微軟正黑體', size=12, bold=True)
_PCF_DATA_FONT = Font(name='微軟正黑體', size=11)
_PCF_TITLE_FONT = Font(name='微軟正黑體', size=14, bold=True)
_PCF_HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_PCF_DATA_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
_PCF_TITLE_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_PCF_PAGE_ALIGN = Alignment(horizontal='right', vertical='center')


# ── 管路補助金額明細表（SDT）常數 ──────────────────────────────────────────────
_SDT_COL_WIDTHS = {
    1: 18,  # A 設施編號
    2: 12,  # B 農戶姓名
    3: 12,  # C 面積（公頃）
    4: 11,  # D 地點
    5: 9,   # E 灌溉型式
    6: 10,  # F 農戶配合款
    7: 12,  # G 末端設施
    8: 10,  # H 水源設施
    9: 10,  # I 調控設施
    10: 12, # J 蓄水池
    11: 10, # K 動力設備
    12: 13, # L 小計
    13: 10, # M 設計費
    14: 12, # N 總計
    15: 12, # O 工程費合計
    16: 10, # P 每公頃補助費
    17: 8,  # Q 百分比（留空）
    18: 12, # R 每公頃總工程費
    # 19: S 設計者（資料萃取保留，不輸出至 Excel）
}
_SDT_THIN_BORDER = Border(left=_XL_S_T, right=_XL_S_T, top=_XL_S_T, bottom=_XL_S_T)
_SDT_FONT_NAME   = '標楷體'                                        # 全表唯一字型來源
_SDT_HDR_FONT    = Font(name=_SDT_FONT_NAME, size=12, bold=True)
_SDT_DATA_FONT   = Font(name=_SDT_FONT_NAME, size=12)
_SDT_TITLE_FONT  = Font(name=_SDT_FONT_NAME, size=20, bold=True)
_SDT_FOOTER_FONT = Font(name=_SDT_FONT_NAME, size=12)
_SDT_FOOTER_FONT_TAG = f'&"{_SDT_FONT_NAME},Regular"'             # 頁尾字型標籤（從 _SDT_FONT_NAME 衍生）
_SDT_CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)
_SDT_RT   = Alignment(horizontal='right',  vertical='center')
_SDT_LT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_SDT_NUM_FMT_INT  = '#,##0'        # 金額欄（整數）
_SDT_NUM_FMT_HA   = '0.000000'     # 面積欄（公頃）
_SDT_NUM_FMT_R    = '#,##0'        # R欄每公頃總工程費（無條件捨去取整數）
_SDT_NUM_FMT_PCT  = '0.00%'        # Q欄百分比（值為 0-1 小數，Excel 自動 ×100 顯示）
_SDT_IRRIGATION_TYPES = ['穿孔管', '噴頭', '滴灌', '微噴', '其它']
_SDT_SIGN_TITLES = ['灌推承辦人', '灌推股長', '灌推主任', '主計室',
                    '主計室股長', '主計室主任', '主任工程師', '副處長', '處長']
_SDT_ROWS_PER_PAGE = 37   # 每頁案件列數（合計區不含在內）
# 55% 縮放下實際可容納：一般頁 ~42 列、含合計的最後頁 ~36 列
# 設 35 保留緩衝；若列印結果有多餘空白或溢出，可在此調整
_SDT_SUMMARY_ROWS  = 6    # 5 灌溉型式合計 + 1 總計（簽核移至頁尾）
_SDT_PAPER_SIZE   = 9   # A4
_SDT_ORIENTATION = 'landscape'
_SDT_SCALE       = 67


def _fmt_ha(value) -> str:
    """面積公頃格式化：截斷至小數第6位，不四捨五入（與 budget_statement_pdf_generator._fmt_ha 相同）"""
    try:
        s = f"{float(value):.10f}"
        dot = s.index('.')
        return s[:dot + 7]
    except (ValueError, TypeError):
        return '0.000000'


def _pcf_row_height(texts_and_widths: list, base: float = 21.0, line_h: float = 18.0) -> float:
    """
    估算 wrap_text 情境下所需的列高。
    texts_and_widths: [(text, col_width_chars), ...]
    中文字元寬度約為 2 個半形單位；line_h 為每行點數。
    """
    max_lines = 1
    for text, cw in texts_and_widths:
        if not text or cw <= 0:
            continue
        # 將全形（中文）字元視為 2 個半形單位
        char_units = sum(2 if ord(ch) > 127 else 1 for ch in text)
        lines = max(1, -(-char_units // max(1, int(cw * 0.9))))  # ceiling div
        max_lines = max(max_lines, lines)
    return max(base, max_lines * line_h)


class ExcelGeneratorService:
    """Excel 文件生成服務 - 基於範本驅動架構生成 .xlsx 檔案"""

    # 範本結構常數定義
    TEMPLATE_HEADER_ROWS = 3      # 標題區塊：第1-3列
    TEMPLATE_DATA_START_ROW = 4   # 資料區塊起始：第4列
    TEMPLATE_DATA_END_ROW = 19    # 資料區塊結束：第19列
    TEMPLATE_PAGE_ROW = 20        # 頁數列：第20列
    TEMPLATE_DATA_ROWS_PER_PAGE = 16  # 每頁資料列數：16列 (4-19)
    TEMPLATE_TOTAL_ROWS_PER_PAGE = 20 # 每頁總列數：20列 (1-20)

    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / "aerc_excel_downloads"
        self.temp_dir.mkdir(exist_ok=True)

    async def generate_photograph_carry_form(self, data: List[Dict[str, Any]], year: str, enable_pagination: bool = True) -> str:
        """
        生成外出拍攝照片攜帶表 Excel 檔案（程式化建構，無範本依賴）

        每個案件依 (鄉鎮, 段名) 聚合土地資料，一組對應一列；
        案件層級欄位（A, B, G, H, I, J, K）跨列合併顯示。

        Args:
            data: 案件資料列表，每筆含 case_number, applicant_name, land_groups,
                  facility_type, irrigation_type, crops_text, phone, address, office_name
            year: 申請年度
            enable_pagination: True 分頁（每頁16列，不拆散案件）；False 不分頁

        Returns:
            str: 生成的 Excel 檔案路徑
        """
        COL_COUNT = 11
        ROWS_PER_PAGE = _PCF_DATA_ROWS_PER_PAGE

        office_name = data[0].get('office_name', '') if data else ''

        # 預先分頁：不拆散案件（一個案件的所有土地列保持在同一頁）
        if enable_pagination:
            pages: List[List[Dict[str, Any]]] = []
            cur_page: List[Dict[str, Any]] = []
            cur_rows = 0
            for item in data:
                n = len(item.get('land_groups') or [{}])
                if cur_rows > 0 and cur_rows + n > ROWS_PER_PAGE:
                    pages.append(cur_page)
                    cur_page = []
                    cur_rows = 0
                cur_page.append(item)
                cur_rows += n
            if cur_page:
                pages.append(cur_page)
            total_pages = len(pages) if pages else 1
        else:
            pages = [data]
            total_pages = 1

        wb = Workbook()
        ws = wb.active

        for col, w in _PCF_COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        def _write_header(title_row: int) -> int:
            """寫標題列 + 欄頭列，回傳第一筆資料列號"""
            ws.row_dimensions[title_row].height = 45.0
            ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=COL_COUNT)
            tc = ws.cell(title_row, 1)
            tc.value = f"農業部農田水利署{office_name}\n外出拍攝照片攜帶表"
            tc.font = _PCF_TITLE_FONT
            tc.alignment = _PCF_TITLE_ALIGN
            tc.border = _PCF_THIN_BORDER

            hdr_row = title_row + 1
            ws.row_dimensions[hdr_row].height = 30.0
            for ci, txt in enumerate(_PCF_HDR_TEXT, start=1):
                c = ws.cell(hdr_row, ci)
                c.value = txt
                c.font = _PCF_HDR_FONT
                c.alignment = _PCF_HDR_ALIGN
                c.border = _PCF_THIN_BORDER
            return hdr_row + 1

        cur_row = 1
        for page_idx, page_items in enumerate(pages):
            page_num = page_idx + 1
            cur_row = _write_header(cur_row)

            for item in page_items:
                land_groups = item.get('land_groups') or [
                    {'land_town': '', 'land_section': '', 'lot_numbers': '', 'facility_area_ha': 0}
                ]
                n = len(land_groups)
                grant_start = cur_row
                grant_end = cur_row + n - 1

                # 土地列欄位（C=3, D=4, E=5, F=6, I=9）：每個聚合組對應一列
                for lg_idx, lg in enumerate(land_groups):
                    row = cur_row + lg_idx
                    lot_numbers_str = str(lg.get('lot_numbers', ''))
                    crops_str = str(lg.get('crops_text', ''))
                    ws.row_dimensions[row].height = _pcf_row_height([
                        (lot_numbers_str, _PCF_COL_WIDTHS[5]),
                        (crops_str,       _PCF_COL_WIDTHS[9]),
                    ])
                    for ci, val in (
                        (3, str(lg.get('land_town', ''))),
                        (4, str(lg.get('land_section', ''))),
                        (5, lot_numbers_str),
                        (6, _fmt_ha(lg.get('facility_area_ha', 0))),
                        (9, crops_str),
                    ):
                        c = ws.cell(row, ci)
                        c.value = val
                        c.font = _PCF_DATA_FONT
                        c.alignment = _PCF_DATA_ALIGN
                        c.border = _PCF_THIN_BORDER

                # 案件層級欄位（A=1, B=2, G=7, H=8, J=10, K=11）：寫入第一列，多列時合併
                for ci, val in (
                    (1, str(item.get('case_number', ''))),
                    (2, str(item.get('applicant_name', ''))),
                    (7, str(item.get('facility_type', ''))),
                    (8, str(item.get('irrigation_type', ''))),
                    (10, str(item.get('phone', ''))),
                    (11, str(item.get('address', ''))),
                ):
                    c = ws.cell(grant_start, ci)
                    c.value = val
                    c.font = _PCF_DATA_FONT
                    c.alignment = _PCF_DATA_ALIGN
                    c.border = _PCF_THIN_BORDER
                    if n > 1:
                        ws.merge_cells(
                            start_row=grant_start, start_column=ci,
                            end_row=grant_end, end_column=ci
                        )

                cur_row += n

            # 頁碼列
            if enable_pagination:
                ws.row_dimensions[cur_row].height = 15.0
                pc = ws.cell(cur_row, COL_COUNT)
                pc.value = f'第{page_num}頁，共{total_pages}頁'
                pc.font = _PCF_DATA_FONT
                pc.alignment = _PCF_PAGE_ALIGN
                cur_row += 1

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"photograph_carry_form_{year}_{timestamp}.xlsx"
        file_path = self.temp_dir / filename
        wb.save(str(file_path))
        return str(file_path)

    async def _fill_template_data(self, workbook, worksheet, data: List[Dict[str, Any]], year: str,
                                  enable_pagination: bool = True) -> str:
        """
        基於範本結構填寫資料 - 範本驅動架構

        範本結構參考（20列為一個完整頁面結構）：
        - 列1-3：標題區塊（機構名稱、年度、表單標題、欄位標題）
        - 列4-19：資料區塊（16列資料格式）
        - 列20：頁數列

        Args:
            workbook: Excel工作簿
            worksheet: 工作表
            data: 資料列表
            year: 年度
            enable_pagination: 分頁模式控制
                - True: 分頁模式 - 每頁16筆資料，複製範本結構到新頁
                - False: 不分頁模式 - 連續填充資料，只保留第一頁標題
        """
        # 根據分頁模式調整邏輯
        if enable_pagination:
            # 分頁模式：每頁資料列數由範本定義
            data_per_page = self.TEMPLATE_DATA_ROWS_PER_PAGE
            total_pages = max(1, (len(data) + data_per_page - 1) // data_per_page)
        else:
            # 不分頁模式：所有資料連續放置
            data_per_page = len(data)
            total_pages = 1

        # 輸出關鍵統計資訊
        print(f"=== Excel 生成統計 ===")
        print(f"分頁模式: {'啟用' if enable_pagination else '停用'}")
        print(f"總資料筆數: {len(data)}")
        print(f"計算總頁數: {total_pages}")
        print(f"===================")

        # 分頁模式：為每頁建立完整的標題組結構
        if enable_pagination:
            # 為第2頁及之後的每一頁建立標題組
            for page_num in range(2, total_pages + 1):
                # 每頁總列數由範本定義
                page_start_row = (page_num - 1) * self.TEMPLATE_TOTAL_ROWS_PER_PAGE + 1

                # 1. 複製主標題列
                self._copy_row_with_format(worksheet, 1, page_start_row)

                # 2. 複製空白列
                worksheet.row_dimensions[page_start_row + 1].height = worksheet.row_dimensions[2].height
                for col in range(1, 12):  # A-K 欄
                    empty_cell = worksheet.cell(row=page_start_row + 1, column=col)
                    empty_cell.value = None
                    empty_cell.border = Border()

                # 3. 複製欄位標題列
                self._copy_row_with_format(worksheet, self.TEMPLATE_HEADER_ROWS, page_start_row + 2)
        else:
            # 不分頁模式：預先清除範本的頁數資訊（避免干擾資料顯示）
            template_page_cell = worksheet.cell(row=self.TEMPLATE_PAGE_ROW, column=11)
            template_page_cell.value = None

        # 依序填入所有資料
        for item_idx, item in enumerate(data):
            if enable_pagination:
                # 分頁模式：計算當前是第幾頁和該頁的第幾筆資料
                current_page = item_idx // data_per_page + 1
                data_index_in_page = item_idx % data_per_page

                # 計算當前資料應該放在哪一列（基於範本結構）
                page_start_row = (current_page - 1) * self.TEMPLATE_TOTAL_ROWS_PER_PAGE
                current_row = page_start_row + self.TEMPLATE_DATA_START_ROW + data_index_in_page

                # 資料定位已計算完成
            else:
                # 不分頁模式：所有資料連續放置，從範本資料起始列開始
                current_row = self.TEMPLATE_DATA_START_ROW + item_idx
                # 資料定位已計算完成

            # 複製資料列格式（從範本資料起始列複製）
            # 注意：第一筆資料直接使用範本起始列，其他資料需要複製格式
            if item_idx > 0:  # 第一筆資料不需複製格式
                self._copy_row_with_format(worksheet, self.TEMPLATE_DATA_START_ROW, current_row)

            # 處理資料
            land_data = item.get('land_data', {})
            total_area = 0
            if isinstance(land_data, dict) and 'land_locations' in land_data:
                for location in land_data['land_locations']:
                    if isinstance(location, dict) and 'area' in location:
                        try:
                            total_area += float(location['area'])
                        except (ValueError, TypeError):
                            continue

            facility_data = item.get('facility_data', {})
            facility_types = []
            if isinstance(facility_data, dict):
                if 'irrigation_type' in facility_data:
                    facility_types.append(str(facility_data['irrigation_type']))
                if 'facility_type' in facility_data:
                    facility_types.append(str(facility_data['facility_type']))
            facility_type_str = ', '.join(facility_types) if facility_types else '未設定'

            # 填寫資料
            row_data = [
                str(item.get('case_number', '')),
                str(item.get('applicant_name', '')),
                '',  # 鄉鎮
                '',  # 段名
                '',  # 地號
                f"{total_area:.4f}" if total_area > 0 else '',
                facility_type_str,
                '',  # 末端型式
                '',  # 農作物
                '',  # 電話
                str(item.get('address', ''))
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                # 設定自動換行
                if cell.alignment:
                    alignment = copy(cell.alignment)
                    alignment.wrap_text = True
                    cell.alignment = alignment
                else:
                    cell.alignment = Alignment(wrap_text=True)

        # 輸出資料填寫結果統計
        print(f"=== 資料填寫結果 ===")
        print(f"總計填寫: {len(data)} 筆資料")
        if enable_pagination:
            print(f"分為 {total_pages} 頁顯示")
        else:
            print("不分頁連續顯示")
        print(f"=====================")

        # 根據分頁模式設定頁碼
        if enable_pagination:
            # 分頁模式：在每頁的頁數列顯示頁碼
            for page_num in range(1, total_pages + 1):
                # 計算頁數列位置：基於範本結構
                page_row = page_num * self.TEMPLATE_TOTAL_ROWS_PER_PAGE

                # 設定頁碼
                page_cell = worksheet.cell(row=page_row, column=11)
                page_cell.value = f'第{page_num}頁，共{total_pages}頁'

                # 複製頁碼格式（從範本頁數列）
                template_page_cell = worksheet.cell(row=self.TEMPLATE_PAGE_ROW, column=11)
                if template_page_cell.font:
                    page_cell.font = copy(template_page_cell.font)
                if template_page_cell.alignment:
                    page_cell.alignment = copy(template_page_cell.alignment)

                # 設定頁數列格式
                worksheet.row_dimensions[page_row].height = 14.3

                # 處理頁數列的邊框
                for col in range(1, 12):  # A-K 欄
                    cell = worksheet.cell(row=page_row, column=col)
                    if col == 11:  # K欄（頁碼欄）
                        if page_num == 1:
                            # 第一頁：移除左、右、下邊框，只保留上邊框
                            cell.border = Border(top=Side(style='thin'))
                        else:
                            # 其他頁：保持原有邊框
                            pass
                    else:
                        # A-J欄：清空內容並移除所有邊框
                        cell.value = None
                        cell.border = Border()
            print(f"頁碼設定完成: {total_pages} 頁")

        # 生成檔案
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"photograph_carry_form_{year}_{timestamp}.xlsx"
        file_path = self.temp_dir / filename

        try:
            workbook.save(str(file_path))
            return str(file_path)
        except Exception as e:
            print(f"Excel save error: {e}")
            print(f"File path: {file_path}")
            raise

    def _copy_row_with_format(self, worksheet, source_row, target_row):
        """複製整列的內容和格式"""
        for col in range(1, 12):  # A-K 欄
            source_cell = worksheet.cell(row=source_row, column=col)
            target_cell = worksheet.cell(row=target_row, column=col)

            # 複製內容和格式
            target_cell.value = source_cell.value
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)

        # 複製行高
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height

        # 如果是第1列（標題列），需要處理合併儲存格
        if source_row == 1:
            # 合併 C、D、E 欄為單一儲存格（財團法人農業工程研究中心）
            merge_range = f'C{target_row}:E{target_row}'
            worksheet.merge_cells(merge_range)

            # 合併 G、H、I 欄為單一儲存格（年度施工照片拍攝攜帶表）
            merge_range = f'G{target_row}:I{target_row}'
            worksheet.merge_cells(merge_range)

    def _remove_top_border(self, worksheet, row_num):
        """移除指定列的上邊框"""
        for col in range(1, 12):  # A-K 欄
            cell = worksheet.cell(row=row_num, column=col)
            if cell.border:
                # 保持其他邊框，只移除上邊框
                new_border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=None,  # 移除上邊框
                    bottom=cell.border.bottom
                )
                cell.border = new_border

    async def generate_a01_execution_progress_report(
        self,
        data: Dict[str, Any],
        year: int
    ) -> str:
        """
        生成 A01 各管理處執行進度報表 Excel 檔案（規則建構，無範本依賴）

        Args:
            data: ExecutionProgressResponse 資料（包含 offices 列表）
            year: 統計年度（民國年）

        Returns:
            str: 生成的 Excel 檔案路徑
        """
        COL_COUNT = 6
        DATA_START = 4

        workbook = Workbook()
        ws = workbook.active

        # 1. 欄寬
        for col, w in _A01_COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # 2. Row 1 標題
        ws.row_dimensions[1].height = _XL_ROW_H_TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
        c = ws.cell(1, 1)
        c.value = f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度各管理處執行進度"
        c.font = _XL_F16
        c.alignment = _XL_A_WRAP

        # 3. Row 2 表號 + 製表日期（E2:F2 合併）
        ws.row_dimensions[2].height = _XL_ROW_H_DATE
        c2 = ws.cell(2, 1)
        c2.value = 'A01'
        c2.font = _XL_F16B
        c2.alignment = _XL_A_CTR
        today = datetime.now()
        date_str = f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日"
        ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=COL_COUNT)
        dc = ws.cell(2, 5)
        dc.value = date_str
        dc.font = _XL_F12
        dc.alignment = _XL_A_RT

        # 4. Row 3 欄頭
        ws.row_dimensions[3].height = _XL_ROW_H_HDR2
        for j in range(COL_COUNT):
            cell = ws.cell(3, j + 1)
            cell.value = _A01_HDR_TEXT[j]
            cell.font = _XL_F14
            cell.alignment = _A01_HDR_ALIGN[j]
            cell.border = _A01_HDR_BORDER[j]

        # 5. 資料列
        offices = data.get('offices', [])
        for idx, office in enumerate(offices):
            row = DATA_START + idx
            ws.row_dimensions[row].height = _XL_ROW_H_DATA
            row_values = [
                office.get('office_name', ''),
                office.get('approved_budget', 0) or 0,
                office.get('completed_cases', 0) or 0,
                float(office.get('total_area', 0) or 0),
                office.get('total_subsidy', 0) or 0,
                float(office.get('execution_rate', 0) or 0),
            ]
            for ci, val in enumerate(row_values):
                cell = ws.cell(row, ci + 1)
                cell.value = val
                cell.font = _XL_F14
                cell.alignment = _XL_A_CTR
                cell.border = _A01_DATA_BORDER[ci]
                cell.number_format = _A01_NUM_FMT[ci + 1]

        # 6. 最末列外框底線（medium bottom）
        if offices:
            last_row = DATA_START + len(offices) - 1
            for ci in range(COL_COUNT):
                cell = ws.cell(last_row, ci + 1)
                db = _A01_DATA_BORDER[ci]
                cell.border = Border(left=db.left, right=db.right, bottom=_XL_S_M)

        # 7. 備註列（資料末列後空一行）
        note_row = DATA_START + len(offices) + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=COL_COUNT)
        note_c = ws.cell(note_row, 1)
        note_c.value = _A01_NOTE_TEXT
        note_c.font = _XL_F14
        note_c.alignment = _XL_A_NOTE
        ws.row_dimensions[note_row].height = _XL_ROW_H_NOTE

        # 8. 儲存後注入備註底線 Rich Text
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = self.temp_dir / f"A01_execution_progress_{year}_{timestamp}.xlsx"
        workbook.save(str(file_path))
        self._inject_note_rich_text(str(file_path), note_row, _A01_NOTE_RUNS)
        return str(file_path)

    # ==================== A02 系列統計報表 ====================

    def _set_cell_value_safe(self, worksheet, cell_ref: str, value):
        """
        安全地設定單元格值（處理合併單元格）

        如果目標單元格是合併單元格的一部分，會自動使用合併區域的左上角單元格
        """
        from openpyxl.utils import coordinate_to_tuple

        row, col = coordinate_to_tuple(cell_ref)

        # 檢查是否在合併單元格中
        for merge_range in worksheet.merged_cells.ranges:
            if (merge_range.min_row <= row <= merge_range.max_row and
                merge_range.min_col <= col <= merge_range.max_col):
                # 使用合併區域的左上角單元格
                worksheet.cell(row=merge_range.min_row, column=merge_range.min_col).value = value
                return

        # 不在合併單元格中，直接賦值
        worksheet[cell_ref] = value

    def _set_cell_value_safe_by_position(self, worksheet, row: int, col: int, value):
        """
        安全地設定單元格值（使用行列位置，處理合併單元格）

        Args:
            worksheet: Excel worksheet
            row: 行號（1-based）
            col: 列號（1-based）
            value: 要設定的值
        """
        from openpyxl.cell.cell import MergedCell

        cell = worksheet.cell(row=row, column=col)

        # 如果是合併單元格，找到合併區域的左上角單元格
        if isinstance(cell, MergedCell):
            for merge_range in worksheet.merged_cells.ranges:
                if (merge_range.min_row <= row <= merge_range.max_row and
                    merge_range.min_col <= col <= merge_range.max_col):
                    worksheet.cell(row=merge_range.min_row, column=merge_range.min_col).value = value
                    return
        else:
            # 不是合併單元格，直接設定
            cell.value = value

    # ── openpyxl 工具組 ──────────────────────────────────────────────────────
    # 以下兩個靜態方法封裝了 openpyxl 合併格的固有行為限制，
    # 供所有動態報表方法複用。

    @staticmethod
    def _tpl_merge_horizontal(
        ws,
        row: int,
        col_start: int,
        col_count: int,
        borders: list,
    ) -> None:
        """
        橫向合併儲存格並保留各欄邊框。

        必要性：openpyxl 的 merge_cells 會保留合併範圍各欄的 border 資料並寫入 XML，
        但若在合併之後才設定邊框，右邊界欄的右外框可能遺失。
        正確做法：先對每欄設定邊框，再執行合併，Excel 即可正確顯示四邊完整框線。

        Args:
            ws: 工作表
            row: 列號
            col_start: 合併起始欄索引
            col_count: 合併欄數
            borders: 長度為 col_count 的邊框列表（對應每欄的邊框）
        """
        for j in range(col_count):
            ws.cell(row, col_start + j).border = borders[j]
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_start + col_count - 1,
        )

    @staticmethod
    def _tpl_force_interior_border(ws, row: int, col: int, border) -> None:
        """
        強制注入邊框到垂直合併格的底端格。

        必要性：openpyxl 的 merge_cells 會從 ws._cells 中刪除非左上角格（底端格），
        導致對底端格設定的邊框在合併後消失。
        Excel 讀取垂直合併格時，底端格若無明確 XML 樣式，該格行高區域的左/下框線不顯示。
        本方法在合併之後建立新的 Cell 物件並直接注入 ws._cells，繞過此限制。

        使用時機：固定欄（縱向合併 HDR1:HDR2）的底端列邊框設定，須在 merge_cells 之後呼叫。

        Args:
            ws: 工作表
            row: 底端格的列號
            col: 底端格的欄號
            border: 要注入的 Border 物件
        """
        from openpyxl.cell.cell import Cell as _OxlCell
        interior = _OxlCell(worksheet=ws, row=row, column=col)
        interior.border = border
        ws._cells[(row, col)] = interior

    def _generate_a02_report(
        self,
        template_name: str,
        col_count: int,
        title_text: str,
        date_text: str,
        rows: List[List[Any]],
        filename_prefix: str,
    ) -> str:
        """
        A02-1/A02-2 報表通用生成邏輯（規則建構，無範本依賴）

        生成結構：
          Row 1: 標題（合併全欄，wrap_text）
          Row 2: 表號 + 製表日期（最後兩欄合併）
          Row 3 (HDR): 各欄頭（單列，高度 45）
          Row 4+: 資料列
          備註列: 純文字備註（合併全欄，與最末列間隔一空列）
        """
        from openpyxl import Workbook
        from openpyxl.styles import Border
        from openpyxl.utils import get_column_letter

        HDR        = 3
        DATA_START = 4
        tbl_num    = template_name.replace('.xlsx', '')

        workbook = Workbook()
        ws       = workbook.active

        # ── 1. 欄寬 ──────────────────────────────────────────────────────
        for i, w in _A0203_COL_WIDTHS[col_count].items():
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        # ── 2. Row 1 標題 ────────────────────────────────────────────────
        ws.row_dimensions[1].height = _XL_ROW_H_TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        c           = ws.cell(1, 1)
        c.value     = title_text
        c.font      = _XL_F16
        c.alignment = _XL_A_WRAP

        # ── 3. Row 2 表號 + 製表日期 ─────────────────────────────────────
        ws.row_dimensions[2].height = _XL_ROW_H_DATE
        c2           = ws.cell(2, 1)
        c2.value     = tbl_num
        c2.font      = _XL_F16B
        c2.alignment = _XL_A_CTR
        ws.merge_cells(start_row=2, start_column=col_count - 1, end_row=2, end_column=col_count)
        dc           = ws.cell(2, col_count - 1)
        dc.value     = date_text
        dc.font      = _XL_F12
        dc.alignment = _XL_A_RT

        # ── 4. Row 3 (HDR) 欄頭 ──────────────────────────────────────────
        ws.row_dimensions[HDR].height = _XL_ROW_H_HDR2  # 45.0
        hdr_texts  = _A0203_HDR_TEXT[col_count]
        hdr_aligns = _A0203_HDR_ALIGN[col_count]
        for j in range(col_count):
            cell           = ws.cell(HDR, j + 1)
            cell.value     = hdr_texts[j]
            cell.font      = _XL_F14
            cell.alignment = hdr_aligns[j]
            cell.border    = _A0203_HDR_BORDER[(col_count, j)]

        # ── 5. 資料列 ─────────────────────────────────────────────────────
        for idx, row_values in enumerate(rows):
            row_num                           = DATA_START + idx
            ws.row_dimensions[row_num].height = _XL_ROW_H_DATA
            for ci, val in enumerate(row_values):
                cell           = ws.cell(row_num, ci + 1)
                cell.value     = val
                cell.font      = _XL_F14
                cell.alignment = _XL_A_CTR
                cell.border    = _A0203_DATA_BORDER[(col_count, ci)]

        # ── 6. 最末列套用外框底線（medium bottom）────────────────────────
        if rows:
            last_row = DATA_START + len(rows) - 1
            for col in range(1, col_count + 1):
                cell = ws.cell(last_row, col)
                b = cell.border
                cell.border = Border(
                    left   = b.left   if b else None,
                    right  = b.right  if b else None,
                    top    = b.top    if b else None,
                    bottom = _XL_S_M,
                )

        # ── 7. 備註列（與最末列間隔一空列，與舊版行為一致）──────────────
        note_row        = DATA_START + len(rows) + 1
        last_col_letter = get_column_letter(col_count)
        ws.merge_cells(f"A{note_row}:{last_col_letter}{note_row}")
        note_c            = ws.cell(note_row, 1)
        note_c.value      = _A0405_NOTE_TEXT
        note_c.font       = _XL_F14
        note_c.alignment  = _XL_A_NOTE
        ws.row_dimensions[note_row].height = _XL_ROW_H_NOTE

        # ── 8. 儲存 ─────────────────────────────────────────────────────
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"{filename_prefix}_{timestamp}.xlsx"
        file_path = self.temp_dir / filename
        workbook.save(str(file_path))
        return str(file_path)

    async def generate_a02_1_report(self, data: Dict[str, Any], year: int) -> str:
        """生成 A02-1 各縣市鄉鎮區統計報表"""
        today = datetime.now()
        rows = []
        for s in data.get('stats', []):
            rows.append([
                s.get('county_name', ''),
                s.get('town_name', ''),
                s.get('completed_cases', 0) or 0,
                float(s.get('total_area', 0) or 0),
                s.get('total_subsidy', 0) or 0,
            ])
        return self._generate_a02_report(
            template_name="A02.xlsx",
            col_count=5,
            title_text=f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度各縣市鄉鎮區統計",
            date_text=f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日",
            rows=rows,
            filename_prefix=f"A02_{year}",
        )

    async def generate_a02_2_report(self, data: Dict[str, Any], year: int) -> str:
        """生成 A02-2 各管理處統計報表"""
        today = datetime.now()
        rows = []
        for s in data.get('stats', []):
            rows.append([
                s.get('office_name', ''),
                s.get('completed_cases', 0) or 0,
                float(s.get('total_area', 0) or 0),
                s.get('total_subsidy', 0) or 0,
            ])
        return self._generate_a02_report(
            template_name="A03.xlsx",
            col_count=4,
            title_text=f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度各管理處統計",
            date_text=f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日",
            rows=rows,
            filename_prefix=f"A03_{year}",
        )

    async def generate_a02_3_report(self, data: Dict[str, Any]) -> str:
        """生成 A02-3 歷年各縣市鄉鎮區統計報表（範本驅動橫向年度展開）"""
        return self._generate_a02_yearly_report(
            template_name="A04.xlsx",
            fixed_cols=2,
            data=data,
            filename_prefix=f"A04_{data.get('start_year', '')}-{data.get('end_year', '')}",
            show_total=False,
        )

    async def generate_a02_4_report(self, data: Dict[str, Any]) -> str:
        """生成 A02-4 歷年各管理處統計報表（範本驅動橫向年度展開）"""
        return self._generate_a02_yearly_report(
            template_name="A05.xlsx",
            fixed_cols=1,
            data=data,
            filename_prefix=f"A05_{data.get('start_year', '')}-{data.get('end_year', '')}",
            show_total=False,
        )

    async def generate_a08_aboriginal_yearly_report(self, data: Dict[str, Any]) -> str:
        """生成 A08 歷年原民區域統計報表（橫向年度展開，與 A04 結構相同）"""
        return self._generate_a02_yearly_report(
            template_name="A08.xlsx",
            fixed_cols=2,
            data=data,
            filename_prefix=f"A08_{data.get('start_year', '')}-{data.get('end_year', '')}",
            show_total=False,
        )

    def _generate_a02_yearly_report(
        self,
        template_name: str,
        fixed_cols: int,
        data: Dict[str, Any],
        filename_prefix: str,
        show_total: bool = True,
    ) -> str:
        """
        A02-3/A02-4 橫向年度展開報表生成（規則建構，無範本依賴）

        生成結構：
          Row 1: 標題（合併 A1:固定欄+3，wrap_text）
          Row 2: 表號 + 製表日期（最後兩欄合併）
          Row 3 (HDR1): 固定欄頭（合併 row3:4）+ N 年度合併欄頭
          Row 4 (HDR2): 指標子欄頭（補助案件數 / 補助面積 / 補助金額）× N
          Row 5+: 資料列（固定識別欄 + N×3 年度指標欄）
          合計列: 合計（show_total=True）
          備註列: 純文字備註（合併全欄）
        """
        from openpyxl import Workbook
        from openpyxl.styles import Border
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import MergedCell
        from decimal import Decimal as _Decimal

        years        = data.get('years', [])
        rows_data    = data.get('rows', [])
        start_year   = data.get('start_year', 0)
        end_year_val = data.get('end_year', 0)
        N            = len(years)
        total_cols   = fixed_cols + N * 3
        # Row 1/2 合併範圍（固定欄 + 1 個年度組），與舊範本版本保持一致
        TPL_TOTAL_COLS = fixed_cols + 3
        HDR1       = 3
        HDR2       = 4
        DATA_START = 5

        # 從 template_name 衍生表號與固定欄頭（無須載入範本檔案）
        tbl_num = template_name.replace('.xlsx', '')
        _FIXED_HDR_NAMES = {2: ['縣市', '鄉鎮區'], 1: ['管理處']}
        _TITLE_TMPLS = {
            'A04.xlsx': '農業部農田水利署\n推廣管路灌溉設施計畫\nOOO年度～OOO年度各縣市鄉鎮區統計表',
            'A05.xlsx': '農業部農田水利署\n推廣管路灌溉設施計畫\nOOO年度～OOO年度各管理處統計表',
            'A08.xlsx': '農業部農田水利署\n推廣管路灌溉設施計畫\nOOO年度～OOO年度原住民地區推動成果統計表',
        }
        title_text      = _TITLE_TMPLS.get(template_name, '').replace(
            'OOO年度～OOO年度', f'{start_year}年度～{end_year_val}年度'
        )
        fixed_hdr_names = _FIXED_HDR_NAMES[fixed_cols]

        workbook = Workbook()
        ws       = workbook.active

        # ── 1. 欄寬 ──────────────────────────────────────────────────────
        cw = _A0405_COL_WIDTHS[fixed_cols]
        for i in range(fixed_cols):
            ws.column_dimensions[get_column_letter(i + 1)].width = cw[i]
        for i in range(N):
            for j in range(3):
                ws.column_dimensions[get_column_letter(fixed_cols + i * 3 + j + 1)].width = cw['year']

        # ── 2. Row 1 標題 ────────────────────────────────────────────────
        ws.row_dimensions[1].height = _XL_ROW_H_TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TPL_TOTAL_COLS)
        c           = ws.cell(1, 1)
        c.value     = title_text
        c.font      = _XL_F16
        c.alignment = _XL_A_WRAP

        # ── 3. Row 2 表號 + 製表日期 ─────────────────────────────────────
        ws.row_dimensions[2].height = _XL_ROW_H_DATE
        c2           = ws.cell(2, 1)
        c2.value     = tbl_num
        c2.font      = _XL_F16B
        c2.alignment = _XL_A_CTR

        today          = datetime.now()
        date_txt       = f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日"
        date_start_col = TPL_TOTAL_COLS - 1
        ws.merge_cells(start_row=2, start_column=date_start_col, end_row=2, end_column=TPL_TOTAL_COLS)
        dc           = ws.cell(2, date_start_col)
        dc.value     = date_txt
        dc.font      = _XL_F12
        dc.alignment = _XL_A_RT

        # ── 4. HDR1 (Row 3) 固定欄頭（合併 row3:4）──────────────────────
        ws.row_dimensions[HDR1].height = _XL_ROW_H_HDR1
        for i in range(fixed_cols):
            col  = i + 1
            ws.merge_cells(start_row=HDR1, start_column=col, end_row=HDR2, end_column=col)
            cell           = ws.cell(HDR1, col)
            cell.value     = fixed_hdr_names[i]
            cell.font      = _XL_F16
            cell.alignment = _XL_A_CTR
            cell.border    = _A0405_FIXED_HDR_BORDER[(fixed_cols, i)]
            # merge_cells 刪除底端格；強制注入邊框確保 HDR2 行高區域的框線顯示
            self._tpl_force_interior_border(ws, HDR2, col, _A0405_FIXED_HDR2_BORDER[(fixed_cols, i)])

        # ── 5. HDR1 (Row 3) 年度合併欄頭 ────────────────────────────────
        for i, year in enumerate(years):
            sc             = fixed_cols + i * 3 + 1
            self._tpl_merge_horizontal(ws, HDR1, sc, 3, _A0405_YEAR_HDR_BORDERS)
            cell           = ws.cell(HDR1, sc)
            cell.value     = f"{year}年度"
            cell.font      = _XL_F16
            cell.alignment = _XL_A_CTR

        # ── 6. HDR2 (Row 4) 指標子欄頭 ──────────────────────────────────
        ws.row_dimensions[HDR2].height = _XL_ROW_H_HDR2
        _subhdr_align = [_XL_A_WRAP, _XL_A_CTR, _XL_A_CTR]
        for i in range(N):
            for j in range(3):
                col            = fixed_cols + i * 3 + j + 1
                cell           = ws.cell(HDR2, col)
                cell.value     = _A0405_SUBHDR_TEXT[j]
                cell.font      = _XL_F14
                cell.alignment = _subhdr_align[j]
                cell.border    = _A0405_SUBHDR_BORDER[j]

        # ── 7. 資料列 ─────────────────────────────────────────────────────
        year_totals = {y: {'cases': 0, 'area': _Decimal('0'), 'subsidy': 0} for y in years}

        for idx, row_item in enumerate(rows_data):
            row_num                           = DATA_START + idx
            ws.row_dimensions[row_num].height = _XL_ROW_H_DATA

            fixed_vals = (
                [row_item.get('county_name', ''), row_item.get('town_name', '')]
                if fixed_cols == 2 else
                [row_item.get('office_name', '')]
            )
            for ci, val in enumerate(fixed_vals):
                cell           = ws.cell(row_num, ci + 1)
                cell.value     = val
                cell.font      = _XL_F14
                cell.alignment = _XL_A_CTR
                cell.border    = _A0405_FIXED_DATA_BORDER[(fixed_cols, ci)]

            for i, ym in enumerate(row_item.get('year_metrics', [])):
                y       = ym.get('year', 0)
                cases   = int(ym.get('completed_cases', 0) or 0)
                area    = float(ym.get('total_area', 0) or 0)
                subsidy = int(ym.get('total_subsidy', 0) or 0)
                sc      = fixed_cols + i * 3 + 1
                for j, val in enumerate([cases, area, subsidy]):
                    cell           = ws.cell(row_num, sc + j)
                    cell.value     = val
                    cell.font      = _XL_F14
                    cell.alignment = _XL_A_CTR
                    cell.border    = _A0405_DATA_YEAR_BORDER[j]
                if y in year_totals:
                    year_totals[y]['cases']   += cases
                    year_totals[y]['area']    += _Decimal(str(area))
                    year_totals[y]['subsidy'] += subsidy

        # ── 8. 合計列（僅 show_total=True 時生成）────────────────────────
        if show_total:
            total_row                           = DATA_START + len(rows_data)
            ws.row_dimensions[total_row].height = _XL_ROW_H_TOTAL
            # 先設邊框再合併，確保右邊界外框不遺失
            for i in range(fixed_cols):
                c           = ws.cell(total_row, i + 1)
                c.font      = _XL_F14
                c.alignment = _XL_A_CTR
                c.border    = _A0405_FIXED_DATA_BORDER[(fixed_cols, i)]
            if fixed_cols == 2:
                ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
            ws.cell(total_row, 1).value = '合  計'
            for i, year in enumerate(years):
                sc  = fixed_cols + i * 3 + 1
                yt  = year_totals[year]
                for j, val in enumerate([yt['cases'], float(yt['area']), yt['subsidy']]):
                    cell           = ws.cell(total_row, sc + j)
                    cell.value     = val
                    cell.font      = _XL_F14
                    cell.alignment = _XL_A_CTR
                    cell.border    = _A0405_DATA_YEAR_BORDER[j]
            note_start_row = total_row + 1
        else:
            note_start_row = DATA_START + len(rows_data)

        # ── 9. 最末列套用外框底線（medium bottom）────────────────────────
        last_table_row = total_row if show_total else DATA_START + len(rows_data) - 1
        if rows_data:
            for col in range(1, total_cols + 1):
                cell = ws.cell(last_table_row, col)
                if isinstance(cell, MergedCell):
                    continue
                b = cell.border
                cell.border = Border(
                    left   = b.left   if b else None,
                    right  = b.right  if b else None,
                    top    = b.top    if b else None,
                    bottom = _XL_S_M,
                )

        # ── 10. 備註列 ───────────────────────────────────────────────────
        note_row        = note_start_row
        last_col_letter = get_column_letter(total_cols)
        ws.merge_cells(f"A{note_row}:{last_col_letter}{note_row}")
        note_c            = ws.cell(note_row, 1)
        note_c.value      = _A0405_NOTE_TEXT
        note_c.font       = _XL_F14
        note_c.alignment  = _XL_A_NOTE
        ws.row_dimensions[note_row].height = _XL_ROW_H_NOTE

        # ── 11. 儲存 ─────────────────────────────────────────────────────
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"{filename_prefix}_{timestamp}.xlsx"
        file_path = self.temp_dir / filename
        workbook.save(str(file_path))
        return str(file_path)

    # ==================== A03 管理處經費統計報表 ====================

    async def generate_a03_budget_analysis_report(
        self,
        data: Dict[str, Any],
        year: int
    ) -> str:
        """
        生成 A03 各管理處經費統計報表 Excel 檔案（規則建構，無範本依賴）

        Args:
            data: BudgetAnalysisResponse 資料（包含 offices、total_* 等欄位）
            year: 統計年度（民國年）

        Returns:
            str: 生成的 Excel 檔案路徑
        """
        DATA_START = 4
        COL_COUNT = 12

        workbook = Workbook()
        ws = workbook.active

        # 1. 欄寬
        for col, w in _A06_COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # 2. Row 1 標題
        ws.row_dimensions[1].height = _XL_ROW_H_TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
        c = ws.cell(1, 1)
        c.value = f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度各管理處經費統計表"
        c.font = _XL_F16
        c.alignment = _XL_A_WRAP

        # 3. Row 2 表號 + 製表日期（K2:L2 合併）
        ws.row_dimensions[2].height = _XL_ROW_H_DATE
        c2 = ws.cell(2, 1)
        c2.value = 'A06'
        c2.font = _XL_F16B
        c2.alignment = _XL_A_CTR
        today = datetime.now()
        date_str = f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日"
        ws.merge_cells(start_row=2, start_column=11, end_row=2, end_column=COL_COUNT)
        dc = ws.cell(2, 11)
        dc.value = date_str
        dc.font = _XL_F12
        dc.alignment = _XL_A_RT

        # 4. Row 3 欄頭
        ws.row_dimensions[3].height = _XL_ROW_H_HDR2
        for j in range(COL_COUNT):
            cell = ws.cell(3, j + 1)
            cell.value = _A06_HDR_TEXT[j]
            cell.font = _XL_F14
            cell.alignment = _XL_A_CTR
            cell.border = _A06_HDR_BORDER[j]

        # 5. 資料列
        offices = data.get('offices', [])
        for idx, office in enumerate(offices):
            row = DATA_START + idx
            ws.row_dimensions[row].height = _XL_ROW_H_DATA
            row_values = [
                office.get('office_name', ''),
                float(office.get('planned_area', 0) or 0),
                office.get('planned_budget', 0) or 0,
                office.get('budgeted_cases', 0) or 0,
                float(office.get('budgeted_area', 0) or 0),
                office.get('budgeted_subsidy', 0) or 0,
                office.get('unbudgeted_subsidy', 0) or 0,
                office.get('verified_cases', 0) or 0,
                float(office.get('verified_area', 0) or 0),
                office.get('verified_amount', 0) or 0,
                float(office.get('area_execution_rate', 0) or 0),
                float(office.get('budget_execution_rate', 0) or 0),
            ]
            for ci, val in enumerate(row_values):
                cell = ws.cell(row, ci + 1)
                cell.value = val
                cell.font = _XL_F14
                cell.alignment = _XL_A_CTR
                cell.border = _A06_DATA_BORDER[ci]
                cell.number_format = _A06_NUM_FMT[ci + 1]

        # 6. 最末列外框底線（medium bottom）
        if offices:
            last_row = DATA_START + len(offices) - 1
            for ci in range(COL_COUNT):
                cell = ws.cell(last_row, ci + 1)
                db = _A06_DATA_BORDER[ci]
                cell.border = Border(left=db.left, right=db.right, bottom=_XL_S_M)

        # 7. 備註列（資料末列後空一行；平文字先寫，儲存後再注入 Rich Text）
        note_row = DATA_START + len(offices) + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=COL_COUNT)
        note_c = ws.cell(note_row, 1)
        note_c.value = _A06_NOTE_TEXT
        note_c.font = _XL_F14
        note_c.alignment = _XL_A_NOTE
        ws.row_dimensions[note_row].height = _A06_ROW_H_NOTE

        # 8. 儲存後注入備註底線 Rich Text
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = self.temp_dir / f"A06_budget_analysis_{year}_{timestamp}.xlsx"
        workbook.save(str(file_path))
        self._inject_note_rich_text(str(file_path), note_row, _A06_NOTE_RUNS)
        return str(file_path)

    @staticmethod
    @staticmethod
    def _inject_note_rich_text(file_path: str, note_row: int, runs: tuple) -> None:
        """
        ZIP 後處理：將備註列的平文字替換為帶底線的 Rich Text（通用）。

        openpyxl 3.1.2 在無 lxml 環境下無法正確寫出 CellRichText，因此先以純文字
        儲存，再直接操作 xlsx ZIP 內的 sheet XML，將 <is><t>...</t></is> 替換成
        帶 <u/> run 的多段格式。

        Args:
            file_path: xlsx 路徑（原址覆蓋）
            note_row:  備註列列號
            runs:      tuple of (underline: bool, text: str)；
                       xml:space="preserve" 依 text 內容自動套用
        """
        import zipfile
        import re
        import os

        _RPR_NORMAL = '<rPr><sz val="14"/><rFont val="微軟正黑體"/></rPr>'
        _RPR_ULINE  = '<rPr><u/><sz val="14"/><rFont val="微軟正黑體"/></rPr>'

        def _run_xml(underline: bool, text: str) -> str:
            rpr = _RPR_ULINE if underline else _RPR_NORMAL
            preserve = '\n' in text or text != text.strip()
            sp = ' xml:space="preserve"' if preserve else ''
            esc = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            return f'<r>{rpr}<t{sp}>{esc}</t></r>'

        rich_is = '<is>' + ''.join(_run_xml(*r) for r in runs) + '</is>'

        with zipfile.ZipFile(file_path, 'r') as zin:
            names = zin.namelist()
            sheet_name = next(n for n in names if n.startswith('xl/worksheets/sheet') and n.endswith('.xml'))
            data = {name: zin.read(name) for name in names}

        sheet_str = data[sheet_name].decode('utf-8')
        pattern = (
            r'(<c\s[^>]*\br="A' + str(note_row) + r'"[^>]*\bt="inlineStr"[^>]*>)'
            r'<is>.*?</is>'
            r'(</c>)'
        )
        new_str, count = re.subn(pattern, r'\1' + rich_is + r'\2', sheet_str, flags=re.DOTALL)
        if count == 0:
            return  # 空資料列情境，找不到備註格，直接略過

        data[sheet_name] = new_str.encode('utf-8')

        tmp_path = file_path + '.tmp'
        with zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(name, data[name])
        os.replace(tmp_path, file_path)

    # ==================== A04 原民區域統計報表 ====================

    async def generate_a04_aboriginal_report(
        self,
        data: Dict[str, Any],
        year: int
    ) -> str:
        """
        生成 A04 原民區域統計報表 Excel 檔案（規則建構，無範本依賴）

        Args:
            data: 原民區域統計資料（包含 stats 列表）
            year: 統計年度（民國年）

        Returns:
            str: 生成的 Excel 檔案路徑
        """
        COL_COUNT = 5
        DATA_START = 4

        workbook = Workbook()
        ws = workbook.active

        # 1. 欄寬（與 A02-1 相同：全部 24.33）
        for ci, w in _A0203_COL_WIDTHS[COL_COUNT].items():
            ws.column_dimensions[get_column_letter(ci + 1)].width = w

        # 2. Row 1 標題
        ws.row_dimensions[1].height = _XL_ROW_H_TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
        c = ws.cell(1, 1)
        c.value = f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度原住民地區推動成果統計表"
        c.font = _XL_F16
        c.alignment = _XL_A_WRAP

        # 3. Row 2 表號 + 製表日期（D2:E2 合併）
        ws.row_dimensions[2].height = _XL_ROW_H_DATE
        c2 = ws.cell(2, 1)
        c2.value = 'A07'
        c2.font = _XL_F16B
        c2.alignment = _XL_A_CTR
        today = datetime.now()
        date_str = f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日"
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=COL_COUNT)
        dc = ws.cell(2, 4)
        dc.value = date_str
        dc.font = _XL_F12
        dc.alignment = _XL_A_RT

        # 4. Row 3 欄頭（與 A02-1 相同）
        ws.row_dimensions[3].height = _XL_ROW_H_HDR2
        for j in range(COL_COUNT):
            cell = ws.cell(3, j + 1)
            cell.value = _A0203_HDR_TEXT[COL_COUNT][j]
            cell.font = _XL_F14
            cell.alignment = _A0203_HDR_ALIGN[COL_COUNT][j]
            cell.border = _A0203_HDR_BORDER[(COL_COUNT, j)]

        # 5. 資料列
        stats = data.get('stats', [])
        for idx, stat in enumerate(stats):
            row = DATA_START + idx
            ws.row_dimensions[row].height = _XL_ROW_H_DATA
            row_values = [
                stat.get('county') or None,
                stat.get('town') or None,
                stat['grant_count'] if stat.get('grant_count') else None,
                stat['subsidy_area'] if stat.get('subsidy_area') else None,
                stat['subsidy_amount'] if stat.get('subsidy_amount') else None,
            ]
            for ci, val in enumerate(row_values):
                cell = ws.cell(row, ci + 1)
                cell.value = val
                cell.font = _XL_F14
                cell.alignment = _XL_A_CTR
                cell.border = _A0203_DATA_BORDER[(COL_COUNT, ci)]
                cell.number_format = _A07_NUM_FMT[ci + 1]

        # 6. 最末列外框底線（medium bottom）
        if stats:
            last_row = DATA_START + len(stats) - 1
            for ci in range(COL_COUNT):
                cell = ws.cell(last_row, ci + 1)
                db = _A0203_DATA_BORDER[(COL_COUNT, ci)]
                cell.border = Border(left=db.left, right=db.right, bottom=_XL_S_M)

        # 7. 備註列（資料末列後空一行）
        note_row = DATA_START + len(stats) + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=COL_COUNT)
        note_c = ws.cell(note_row, 1)
        note_c.value = _A07_NOTE_TEXT
        note_c.font = _XL_F14
        note_c.alignment = _XL_A_NOTE
        ws.row_dimensions[note_row].height = _XL_ROW_H_NOTE

        # 8. 儲存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = self.temp_dir / f"A07_aboriginal_{year}_{timestamp}.xlsx"
        workbook.save(str(file_path))
        return str(file_path)

    # ==================== B01 系列推動成果統計報表（管理區內外分組） ====================

    def _generate_a09_a10_report(
        self,
        table_code: str,
        label_col_name: str,
        year: int,
        is_current_year: bool,
        rows: List[List[Any]],
    ) -> str:
        """
        A09/A10 通用報表生成引擎（規則建構，無範本依賴）

        生成結構：
          Row 1: 標題（合併全欄，wrap_text）
          Row 2: 表號 + 製表日期
          Row 3 (HDR1): label 欄垂直合併 rows 3-5；主分組（已編列/已結案/小計）
          Row 4 (HDR2): 事業區域外/事業區域內 子分組
          Row 5 (HDR3): 案件數/面積(公頃)/金額(元) 指標欄
          Row 6+: 資料列
          備註列: 純文字備註（合併全欄）

        Args:
            table_code: 'A09' 或 'A10'
            label_col_name: 標籤欄名稱（'縣 市' 或 '管理處'）
            year: 統計年度（民國年）
            is_current_year: 當年度（16欄）或非當年度（7欄）
            rows: 資料列，每列含 col_count 個值（含 label）

        Returns:
            str: 生成的 Excel 檔案路徑
        """
        from openpyxl.cell.cell import MergedCell
        from datetime import timezone

        col_count  = 16 if is_current_year else 7
        DATA_START = 6
        last_col_letter = get_column_letter(col_count)

        workbook = Workbook()
        ws = workbook.active

        # ── 1. 欄寬 ──────────────────────────────────────────────────────
        for col, w in _A0910_COL_WIDTHS[col_count].items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # ── 2. Row 1 標題 ────────────────────────────────────────────────
        ws.row_dimensions[1].height = _XL_ROW_H_TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        c = ws.cell(1, 1)
        c.value = (
            f"農業部農田水利署\n推廣管路灌溉設施計畫\n"
            f"{year}年度各{label_col_name.strip()}事業區域內外推動成果統計表"
        )
        c.font = _XL_F16
        c.alignment = _XL_A_WRAP

        # ── 3. Row 2 表號 + 製表日期 ─────────────────────────────────────
        ws.row_dimensions[2].height = _XL_ROW_H_DATE
        c2 = ws.cell(2, 1)
        c2.value = table_code
        c2.font = _XL_F16B
        c2.alignment = _XL_A_CTR
        today = datetime.now(timezone.utc)
        date_str = f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日"
        ws.merge_cells(start_row=2, start_column=col_count - 1, end_row=2, end_column=col_count)
        dc = ws.cell(2, col_count - 1)
        dc.value = date_str
        dc.font = _XL_F12
        dc.alignment = _XL_A_RT

        # ── 4. 三層表頭結構定義 ───────────────────────────────────────────
        if is_current_year:
            hdr1_groups = [
                (2,  7,  '已編列', _A0910_FILL_BUDGETED),
                (8,  13, '已結案', _A0910_FILL_COMPLETED),
                (14, 16, '小計',   _A0910_FILL_SUBTOTAL),
            ]
            hdr2_groups = [
                (2,  4,  '事業區域外', _A0910_FILL_BUDGETED,  True),
                (5,  7,  '事業區域內', _A0910_FILL_BUDGETED,  False),
                (8,  10, '事業區域外', _A0910_FILL_COMPLETED, True),
                (11, 13, '事業區域內', _A0910_FILL_COMPLETED, False),
                (14, 16, '小計',      _A0910_FILL_SUBTOTAL,  True),
            ]
            major_starts = {2, 8, 14}
        else:
            hdr1_groups = [
                (2, 7, '已結案', _A0910_FILL_COMPLETED),
            ]
            hdr2_groups = [
                (2, 4, '事業區域外', _A0910_FILL_COMPLETED, True),
                (5, 7, '事業區域內', _A0910_FILL_COMPLETED, False),
            ]
            major_starts = {2}

        for r in range(3, 6):
            ws.row_dimensions[r].height = _XL_ROW_H_HDR2

        # ── 5. Row 3-5 Col 1：垂直合併，顯示 label_col_name ─────────────
        # 先設邊框（左中/右中/上中），再合併（合併後底端格用 _tpl_force_interior_border 注入）
        for r in (3, 4, 5):
            ws.cell(r, 1).border = Border(left=_XL_S_M, right=_XL_S_M, top=_XL_S_M)
        ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
        label_cell = ws.cell(3, 1)
        label_cell.value = label_col_name
        label_cell.font = _XL_F14
        label_cell.alignment = _XL_A_CTR
        label_cell.border = Border(left=_XL_S_M, right=_XL_S_M, top=_XL_S_M, bottom=_XL_S_M)
        # 強制注入底端格邊框
        self._tpl_force_interior_border(ws, 4, 1, Border(left=_XL_S_M, right=_XL_S_M))
        self._tpl_force_interior_border(ws, 5, 1, Border(left=_XL_S_M, right=_XL_S_M, bottom=_XL_S_M))

        # ── 6. Row 3 (HDR1)：主分組合併欄頭 ─────────────────────────────
        for (sc, ec, text, fill) in hdr1_groups:
            # 先設每格邊框再合併（_tpl_merge_horizontal 模式）
            for col in range(sc, ec + 1):
                is_left  = (col == sc)
                is_right = (col == ec)
                ws.cell(3, col).border = Border(
                    left   = _XL_S_M if is_left else None,
                    right  = _XL_S_M if is_right else None,
                    top    = _XL_S_M,
                    bottom = _XL_S_T,
                )
            ws.merge_cells(start_row=3, start_column=sc, end_row=3, end_column=ec)
            cell = ws.cell(3, sc)
            cell.value = text
            cell.font = _XL_F14
            cell.alignment = _XL_A_CTR
            cell.fill = fill

        # ── 7. Row 4 (HDR2)：子分組合併欄頭 ─────────────────────────────
        for (sc, ec, text, fill, is_major) in hdr2_groups:
            left_side  = _XL_S_M if is_major else _XL_S_T
            # 右側：若此子組緊接另一主分組，則 medium；否則 thin（非最後欄）
            right_col  = ec + 1
            is_last_col = (ec == col_count)
            right_side = _XL_S_M if (is_last_col or right_col in major_starts) else _XL_S_T
            for col in range(sc, ec + 1):
                is_col_left  = (col == sc)
                is_col_right = (col == ec)
                ws.cell(4, col).border = Border(
                    left   = left_side if is_col_left else None,
                    right  = right_side if is_col_right else None,
                    top    = _XL_S_T,
                    bottom = _XL_S_T,
                )
            ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)
            cell = ws.cell(4, sc)
            cell.value = text
            cell.font = _XL_F14
            cell.alignment = _XL_A_CTR
            cell.fill = fill

        # ── 8. Row 5 (HDR3)：指標子欄頭（不合併，每格獨立）────────────
        hdr3_texts = ['案件數', '面積\n(公頃)', '金額\n(元)']
        for col in range(2, col_count + 1):
            group_pos = (col - 2) % 3   # 0=案件數, 1=面積, 2=金額
            is_major_left  = col in major_starts
            left_side  = _XL_S_M if is_major_left else (_XL_S_T if group_pos == 0 else None)
            is_last_col    = (col == col_count)
            right_side = _XL_S_M if (is_last_col or col + 1 in major_starts) else (
                         _XL_S_M if group_pos == 2 else None)
            cell = ws.cell(5, col)
            cell.value = hdr3_texts[group_pos]
            cell.font = _XL_F14
            cell.alignment = _XL_A_WRAP
            cell.border = Border(
                left   = left_side,
                right  = right_side,
                top    = _XL_S_T,
                bottom = _XL_S_M,
            )

        # ── 9. 資料列 ─────────────────────────────────────────────────────
        for idx, row_data in enumerate(rows):
            row_num = DATA_START + idx
            ws.row_dimensions[row_num].height = _XL_ROW_H_DATA
            for col_i, val in enumerate(row_data):
                col_num = col_i + 1
                cell = ws.cell(row_num, col_num, value=val)
                cell.font = _XL_F14
                cell.alignment = _XL_A_CTR
                if col_num == 1:
                    cell.border = Border(left=_XL_S_M, right=_XL_S_M, bottom=_XL_S_T)
                else:
                    group_pos  = (col_num - 2) % 3
                    is_maj_l   = col_num in major_starts
                    left_side  = _XL_S_M if is_maj_l else (_XL_S_T if group_pos == 0 else None)
                    is_last    = (col_num == col_count)
                    right_side = _XL_S_M if (is_last or col_num + 1 in major_starts) else (
                                 _XL_S_M if group_pos == 2 else None)
                    cell.border = Border(left=left_side, right=right_side, bottom=_XL_S_T)
                    cell.number_format = _A0910_NUM_FMT[group_pos]

        # ── 10. 最末列外框底線（medium bottom）───────────────────────────
        if rows:
            last_data_row = DATA_START + len(rows) - 1
            for col_num in range(1, col_count + 1):
                cell = ws.cell(last_data_row, col_num)
                if isinstance(cell, MergedCell):
                    continue
                b = cell.border
                cell.border = Border(
                    left   = b.left   if b else None,
                    right  = b.right  if b else None,
                    top    = b.top    if b else None,
                    bottom = _XL_S_M,
                )

        # ── 11. 備註列（與末列間隔一空列）───────────────────────────────
        note_row = DATA_START + len(rows) + 1
        ws.merge_cells(f"A{note_row}:{last_col_letter}{note_row}")
        note_c = ws.cell(note_row, 1)
        note_c.value = _A0910_NOTE_TEXT
        note_c.font = _XL_F14
        note_c.alignment = _XL_A_NOTE
        ws.row_dimensions[note_row].height = _A0910_ROW_H_NOTE

        # ── 12. 儲存 ──────────────────────────────────────────────────────
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_path = self.temp_dir / f"{table_code}_{year}_{timestamp}.xlsx"
        workbook.save(str(file_path))
        return str(file_path)

    def _generate_b01_report(
        self,
        template_name: str,
        col_count: int,
        title_text: str,
        date_text: str,
        rows: List[List[Any]],
        filename_prefix: str,
    ) -> str:
        """
        B01 報表通用生成邏輯（範本驅動 + 動態增長）

        架構與 A02 完全一致，但欄位數量不同：
        - B01-1/B01-3: 7欄（縣市 + 管理區內3指標 + 管理區外3指標）
        - B01-2/B01-4: 7欄（管理處 + 管理區內3指標 + 管理區外3指標）
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Border
        from openpyxl.cell.cell import MergedCell

        template_path = settings.get_template_path(template_name)
        if not template_path.exists():
            raise FileNotFoundError(f"範本檔案不存在: {template_path}")

        workbook = load_workbook(str(template_path))
        worksheet = workbook.active

        # 移除 Print Area 定義以避免警告
        if 'Print_Area' in workbook.defined_names:
            del workbook.defined_names['Print_Area']

        DATA_START_ROW = 4
        HEADER_ROW = 3

        # 1. 從範本擷取樣式參考（Row 4）
        col_styles = {}
        frame_bottom_sides = {}
        for col in range(1, col_count + 1):
            ref_cell = worksheet.cell(row=DATA_START_ROW, column=col)
            col_styles[col] = {
                'font': copy(ref_cell.font) if ref_cell.font else None,
                'alignment': copy(ref_cell.alignment) if ref_cell.alignment else None,
                'border': copy(ref_cell.border) if ref_cell.border else None,
                'fill': copy(ref_cell.fill) if ref_cell.fill else None,
                'number_format': ref_cell.number_format,
            }
            # 表頭底部邊框
            header_cell = worksheet.cell(row=HEADER_ROW, column=col)
            if (header_cell.border and header_cell.border.bottom
                    and getattr(header_cell.border.bottom, 'style', None)):
                frame_bottom_sides[col] = header_cell.border.bottom

        # 2. 標題與日期（第1列與第2列）
        worksheet['A1'] = title_text
        last_col_letter = get_column_letter(col_count)
        self._set_cell_value_safe(worksheet, f'{last_col_letter}2', date_text)

        # 3. 清除範例資料（保留 Row 3 表頭和 Row 4 樣式參考）
        max_row = worksheet.max_row
        for row in range(DATA_START_ROW, max_row + 1):
            for col in range(1, col_count + 1):
                cell = worksheet.cell(row=row, column=col)
                # 跳過合併單元格（MergedCell 的 value 是唯讀的）
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.border = Border()
            if row in worksheet.row_dimensions:
                del worksheet.row_dimensions[row]

        # 4. 動態新增資料列 + 應用範本樣式
        for row_idx, row_data in enumerate(rows, start=DATA_START_ROW):
            for col_idx, value in enumerate(row_data, start=1):
                if col_idx > col_count:
                    continue
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                # 應用範本樣式
                style = col_styles.get(col_idx, {})
                if style.get('font'):
                    cell.font = copy(style['font'])
                if style.get('alignment'):
                    cell.alignment = copy(style['alignment'])
                if style.get('border'):
                    cell.border = copy(style['border'])
                if style.get('fill'):
                    cell.fill = copy(style['fill'])
                if style.get('number_format'):
                    cell.number_format = style['number_format']

        # 5. 小計列
        last_data_row = DATA_START_ROW + len(rows) - 1
        subtotal_row = last_data_row + 1
        self._set_cell_value_safe_by_position(worksheet, subtotal_row, 1, "小計")

        # 應用小計列樣式並計算總和公式
        for col in range(1, col_count + 1):
            cell = worksheet.cell(row=subtotal_row, column=col)
            style = col_styles.get(col, {})
            if style.get('font'):
                cell.font = copy(style['font'])
            if style.get('alignment'):
                cell.alignment = copy(style['alignment'])
            if style.get('fill'):
                cell.fill = copy(style['fill'])

            # 數字欄位加總公式（第2欄起為數字）
            if col > 1:
                col_letter = get_column_letter(col)
                sum_formula = f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{last_data_row})"
                self._set_cell_value_safe_by_position(worksheet, subtotal_row, col, sum_formula)
                if style.get('number_format'):
                    cell.number_format = style['number_format']

            # 底部邊框（表格外框）
            if col in frame_bottom_sides:
                current_border = copy(cell.border) if cell.border else Border()
                cell.border = Border(
                    left=current_border.left,
                    right=current_border.right,
                    top=current_border.top,
                    bottom=frame_bottom_sides[col]
                )

        # 6. 備註（緊跟小計列之後）
        note_row = subtotal_row + 1
        self._set_cell_value_safe_by_position(worksheet, note_row, 1, "註：案件數為有效之有案號案件數")

        # 儲存檔案
        output_filename = f"{filename_prefix}.xlsx"
        output_path = self.temp_dir / output_filename
        workbook.save(str(output_path))
        return str(output_path)

    async def generate_a09_report(self, data: Dict[str, Any], year: int) -> str:
        """生成 A09 各縣市事業區域內外推動成果統計表"""
        is_current_year: bool = data.get('is_current_year', False)
        rows = []
        for s in data.get('stats', []):
            boc = s.get('budgeted_outside_cases', 0) or 0
            boa = float(s.get('budgeted_outside_area', 0) or 0)
            bos = s.get('budgeted_outside_subsidy', 0) or 0
            bic = s.get('budgeted_inside_cases', 0) or 0
            bia = float(s.get('budgeted_inside_area', 0) or 0)
            bis_ = s.get('budgeted_inside_subsidy', 0) or 0
            coc = s.get('completed_outside_cases', 0) or 0
            coa = float(s.get('completed_outside_area', 0) or 0)
            cos_ = s.get('completed_outside_subsidy', 0) or 0
            cic = s.get('completed_inside_cases', 0) or 0
            cia = float(s.get('completed_inside_area', 0) or 0)
            cis = s.get('completed_inside_subsidy', 0) or 0
            row: list = [s.get('county_name', '')]
            if is_current_year:
                row += [boc, boa, bos, bic, bia, bis_]
            row += [coc, coa, cos_, cic, cia, cis]
            if is_current_year:
                row += [boc + bic + coc + cic, boa + bia + coa + cia, bos + bis_ + cos_ + cis]
            rows.append(row)
        return self._generate_a09_a10_report('A09', '縣 市', year, is_current_year, rows)

    async def generate_a10_report(self, data: Dict[str, Any], year: int) -> str:
        """生成 A10 各管理處事業區域內外推動成果統計表"""
        is_current_year: bool = data.get('is_current_year', False)
        rows = []
        for s in data.get('stats', []):
            boc = s.get('budgeted_outside_cases', 0) or 0
            boa = float(s.get('budgeted_outside_area', 0) or 0)
            bos = s.get('budgeted_outside_subsidy', 0) or 0
            bic = s.get('budgeted_inside_cases', 0) or 0
            bia = float(s.get('budgeted_inside_area', 0) or 0)
            bis_ = s.get('budgeted_inside_subsidy', 0) or 0
            coc = s.get('completed_outside_cases', 0) or 0
            coa = float(s.get('completed_outside_area', 0) or 0)
            cos_ = s.get('completed_outside_subsidy', 0) or 0
            cic = s.get('completed_inside_cases', 0) or 0
            cia = float(s.get('completed_inside_area', 0) or 0)
            cis = s.get('completed_inside_subsidy', 0) or 0
            row: list = [s.get('office_name', '')]
            if is_current_year:
                row += [boc, boa, bos, bic, bia, bis_]
            row += [coc, coa, cos_, cic, cia, cis]
            if is_current_year:
                row += [boc + bic + coc + cic, boa + bia + coa + cia, bos + bis_ + cos_ + cis]
            rows.append(row)
        return self._generate_a09_a10_report('A10', '管理處', year, is_current_year, rows)

    async def generate_b01_1_report(self, data: Dict[str, Any], year: int) -> str:
        """生成 B01-1 各縣市管理區內外統計報表（單年度）"""
        today = datetime.now()
        rows = []
        for s in data.get('stats', []):
            rows.append([
                s.get('county_name', ''),
                # 管理區內
                s.get('inside_cases', 0) or 0,
                float(s.get('inside_area', 0) or 0),
                s.get('inside_subsidy', 0) or 0,
                # 管理區外
                s.get('outside_cases', 0) or 0,
                float(s.get('outside_area', 0) or 0),
                s.get('outside_subsidy', 0) or 0,
            ])
        return self._generate_b01_report(
            template_name="B01-1.xlsx",
            col_count=7,
            title_text=f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度各縣市推動成果統計表",
            date_text=f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日",
            rows=rows,
            filename_prefix=f"B01-1_{year}",
        )

    async def generate_b01_2_report(self, data: Dict[str, Any], year: int) -> str:
        """生成 B01-2 各管理處管理區內外統計報表（單年度）"""
        today = datetime.now()
        rows = []
        for s in data.get('stats', []):
            rows.append([
                s.get('office_name', ''),
                # 管理區內
                s.get('inside_cases', 0) or 0,
                float(s.get('inside_area', 0) or 0),
                s.get('inside_subsidy', 0) or 0,
                # 管理區外
                s.get('outside_cases', 0) or 0,
                float(s.get('outside_area', 0) or 0),
                s.get('outside_subsidy', 0) or 0,
            ])
        return self._generate_b01_report(
            template_name="B01-2.xlsx",
            col_count=7,
            title_text=f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度各管理處推動成果統計表",
            date_text=f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日",
            rows=rows,
            filename_prefix=f"B01-2_{year}",
        )

    async def generate_b01_3_report(self, data: Dict[str, Any], start_year: int, end_year: int) -> str:
        """生成 B01-3 歷年各縣市管理區內外統計報表"""
        today = datetime.now()
        rows = []
        for s in data.get('stats', []):
            rows.append([
                s.get('county_name', ''),
                # 管理區內
                s.get('inside_cases', 0) or 0,
                float(s.get('inside_area', 0) or 0),
                s.get('inside_subsidy', 0) or 0,
                # 管理區外
                s.get('outside_cases', 0) or 0,
                float(s.get('outside_area', 0) or 0),
                s.get('outside_subsidy', 0) or 0,
            ])
        return self._generate_b01_report(
            template_name="B01-3.xlsx",
            col_count=7,
            title_text=f"農業部農田水利署\n推廣管路灌溉設施計畫\n{start_year}年度～{end_year}年度各縣市推動成果統計表",
            date_text=f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日",
            rows=rows,
            filename_prefix=f"B01-3_{start_year}-{end_year}",
        )

    async def generate_b01_4_report(self, data: Dict[str, Any], start_year: int, end_year: int) -> str:
        """生成 B01-4 歷年各管理處管理區內外統計報表"""
        today = datetime.now()
        rows = []
        for s in data.get('stats', []):
            rows.append([
                s.get('office_name', ''),
                # 管理區內
                s.get('inside_cases', 0) or 0,
                float(s.get('inside_area', 0) or 0),
                s.get('inside_subsidy', 0) or 0,
                # 管理區外
                s.get('outside_cases', 0) or 0,
                float(s.get('outside_area', 0) or 0),
                s.get('outside_subsidy', 0) or 0,
            ])
        return self._generate_b01_report(
            template_name="B01-4.xlsx",
            col_count=7,
            title_text=f"農業部農田水利署\n推廣管路灌溉設施計畫\n{start_year}年度～{end_year}年度各管理處推動成果統計表",
            date_text=f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日",
            rows=rows,
            filename_prefix=f"B01-4_{start_year}-{end_year}",
        )

    async def generate_b03_report(self, data: Dict[str, Any], year: int) -> str:
        """
        生成 B03 各縣市鄉鎮區各類補助項目統計表 Excel（規則建構，19欄 4列標題）

        Args:
            data: B03StatsResponse.dict() 資料
            year: 統計年度（民國年）

        Returns:
            str: 生成的 Excel 檔案路徑
        """
        COL_COUNT = 19
        DATA_START = 5  # Row 1 主標題, Row 2 代號+日期, Row 3 父標題, Row 4 子標題

        workbook = Workbook()
        ws = workbook.active

        # 1. 欄寬
        for col, w in _B03_COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # 2. Row 1 主標題（合併 A1:S1）
        ws.row_dimensions[1].height = _XL_ROW_H_TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
        c1 = ws.cell(1, 1)
        c1.value = f"農業部農田水利署\n推廣管路灌溉設施計畫\n{year}年度各縣市鄉鎮區各類補助項目統計表"
        c1.font = _XL_F16
        c1.alignment = _XL_A_WRAP

        # 3. Row 2 代號 + 製表日期
        ws.row_dimensions[2].height = _XL_ROW_H_DATE
        c2 = ws.cell(2, 1)
        c2.value = 'B03'
        c2.font = _XL_F16B  # type: ignore[attr-defined]
        c2.alignment = _XL_A_CTR
        today = datetime.now()
        date_str = f"製表日期：{today.year - 1911}年{today.month:02d}月{today.day:02d}日"
        ws.merge_cells(start_row=2, start_column=16, end_row=2, end_column=COL_COUNT)
        dc = ws.cell(2, 16)
        dc.value = date_str
        dc.font = _XL_F12
        dc.alignment = _XL_A_RT

        # 4. Row 3 父標題層 指定高度，不共用
        ws.row_dimensions[3].height = 45

        def _mhdr(col, end_col, end_row, text, wrap=False):
            """合併並寫入父標題格，並補齊合併範圍邊緣格的 border"""
            if end_col > col or end_row > 3:
                ws.merge_cells(start_row=3, start_column=col, end_row=end_row, end_column=end_col)
            is_left = col == 1
            is_right = end_col == COL_COUNT
            btm3 = _XL_S_T if end_row == 3 else _XL_S_M
            # Row 3：整行補 top border；首/末欄設 left/right
            for ci in range(col, end_col + 1):
                ws.cell(3, ci).border = Border(
                    top=_XL_S_M,
                    bottom=btm3,
                    left=(_XL_S_M if is_left else _XL_S_T) if ci == col else None,
                    right=(_XL_S_M if is_right else _XL_S_T) if ci == end_col else None,
                )
            # Row 4（跨行）：整行補 bottom border；首/末欄設 left/right
            if end_row > 3:
                for ci in range(col, end_col + 1):
                    ws.cell(end_row, ci).border = Border(
                        bottom=_XL_S_M,
                        left=(_XL_S_M if is_left else _XL_S_T) if ci == col else None,
                        right=(_XL_S_M if is_right else _XL_S_T) if ci == end_col else None,
                    )
            # 值和樣式只寫在 top-left 格
            c = ws.cell(3, col)
            c.value = text
            c.font = _XL_F14
            c.alignment = _XL_A_WRAP if wrap else _XL_A_CTR

        _mhdr(1, 2, 3, '地點')                                      # A3:B3
        _mhdr(3, 3, 4, '灌溉型式')                                   # C3:C4
        _mhdr(4, 4, 4, '補助面積\n(公頃)', wrap=True)                 # D3:D4
        _mhdr(5, 5, 4, '補助案件數\n(已結案)', wrap=True)             # E3:E4
        _mhdr(6, 6, 4, '農戶配合款\n(元)', wrap=True)                 # F3:F4
        _mhdr(7, 12, 3, '補助經費(元)')                               # G3:L3
        _mhdr(13, 13, 4, '工程經費\n合計(元)', wrap=True)             # M3:M4
        _mhdr(14, 15, 3, '調蓄設施蓄水池*')                           # N3:O3
        _mhdr(16, 16, 3, '動力設施')                                    # P3 only
        _mhdr(17, 19, 3, '每公頃工程單價\n(田間管路設施部分)', wrap=True)  # Q3:S3

        # 5. Row 4 子標題層
        ws.row_dimensions[4].height = _XL_ROW_H_HDR2
        sub_headers = {
            1: '縣市', 2: '鄉鎮區',
            7: '田間管路\n設施(元)', 8: '調蓄\n設施(元)', 9: '動力\n設施(元)',
            10: '調控\n設備(元)', 11: '設計費\n(元)', 12: '總計(元)',
            14: '噸', 15: '座', 16: '抽水機\n(臺)',
            17: '補助款\n(元)', 18: '百分比%', 19: '總工程費\n(元)',
        }
        for col, text in sub_headers.items():
            c = ws.cell(4, col)
            c.value = text
            c.font = _XL_F14
            c.alignment = _XL_A_WRAP
            c.border = Border(
                left=_XL_S_M if col == 1 else _XL_S_T,
                right=_XL_S_M if col == COL_COUNT else _XL_S_T,
                top=_XL_S_T, bottom=_XL_S_M
            )

        # 6. 資料列
        rows = data.get('rows', [])
        for idx, row in enumerate(rows):
            r = DATA_START + idx
            ws.row_dimensions[r].height = _XL_ROW_H_DATA
            vals = [
                row.get('county_name', ''),
                row.get('town_name', ''),
                row.get('irrigation_type', ''),
                float(row.get('total_area', 0) or 0),
                int(row.get('completed_cases', 0) or 0),
                int(row.get('farmer_contribution', 0) or 0),
                int(row.get('pipeline_subsidy', 0) or 0),
                int(row.get('storage_subsidy', 0) or 0),
                int(row.get('power_subsidy', 0) or 0),
                int(row.get('control_subsidy', 0) or 0),
                int(row.get('design_fee_subsidy', 0) or 0),
                int(row.get('total_subsidy', 0) or 0),
                int(row.get('total_engineering', 0) or 0),
                int(row.get('storage_tonnage', 0) or 0),
                int(row.get('storage_count', 0) or 0),
                int(row.get('pump_count', 0) or 0),
                int(row.get('subsidy_per_ha', 0) or 0),
                float(row.get('pipeline_ratio', 0) or 0),
                int(row.get('engineering_per_ha', 0) or 0),
            ]
            for ci, val in enumerate(vals):
                cell = ws.cell(r, ci + 1)
                cell.value = val
                cell.font = _XL_F14
                cell.alignment = _XL_A_CTR
                cell.number_format = _B03_NUM_FMT[ci + 1]
                cell.border = Border(
                    left=_XL_S_M if ci == 0 else _XL_S_T,
                    right=_XL_S_M if ci == COL_COUNT - 1 else _XL_S_T,
                    bottom=_XL_S_T
                )

        # 7. 全表合計列
        total_row = DATA_START + len(rows)
        ws.row_dimensions[total_row].height = _XL_ROW_H_TOTAL
        total_vals = [
            '合計', '',
            '',
            float(data.get('total_area', 0) or 0),
            int(data.get('total_cases', 0) or 0),
            int(data.get('total_farmer_contribution', 0) or 0),
            int(data.get('total_pipeline_subsidy', 0) or 0),
            int(data.get('total_storage_subsidy', 0) or 0),
            int(data.get('total_power_subsidy', 0) or 0),
            int(data.get('total_control_subsidy', 0) or 0),
            int(data.get('total_design_fee_subsidy', 0) or 0),
            int(data.get('total_subsidy', 0) or 0),
            int(data.get('total_engineering', 0) or 0),
            int(data.get('total_storage_tonnage', 0) or 0),
            int(data.get('total_storage_count', 0) or 0),
            int(data.get('total_pump_count', 0) or 0),
            '', '', '',  # per-ha 合計欄不顯示
        ]
        for ci, val in enumerate(total_vals):
            cell = ws.cell(total_row, ci + 1)
            cell.value = val
            cell.font = _XL_F14
            cell.alignment = _XL_A_CTR
            cell.number_format = _B03_NUM_FMT[ci + 1]
            cell.border = Border(
                left=_XL_S_M if ci == 0 else _XL_S_T,
                right=_XL_S_M if ci == COL_COUNT - 1 else _XL_S_T,
                bottom=_XL_S_M
            )

        # 8. 備註列
        note_row = total_row + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=COL_COUNT)
        note_c = ws.cell(note_row, 1)
        note_c.value = _B03_NOTE_TEXT
        note_c.font = _XL_F14
        note_c.alignment = _XL_A_NOTE
        ws.row_dimensions[note_row].height = _XL_ROW_H_NOTE

        # 9. 儲存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = self.temp_dir / f"B03_county_town_subsidy_{year}_{timestamp}.xlsx"
        workbook.save(str(file_path))
        return str(file_path)

    async def generate_address_labels(self, data: List[Dict[str, Any]], year: str) -> str:
        """
        生成住址標籤 Excel 檔案

        每行並排 2 個標籤（左欄 A-E，右欄 F-J），每組標籤佔 4 行：
        - 行 +0（高 39.75）：郵遞地址（郵遞區號 + 縣市 + 鄉鎮 + 村里 + 詳細地址）
        - 行 +1（高 24.00）：案件編號 + 申請人姓名
        - 行 +2（高 33.75）：空行
        - 行 +3（高  9.95）：間隔行

        地址格式與工程預算書通訊住址完全一致。

        Args:
            data: 案件資料列表，每筆需含 case_number、applicant_name、county、town、village、address
            year: 申請年度，用於檔名

        Returns:
            str: 生成的 Excel 檔案絕對路徑
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "地址"

        # 欄寬（依照參考範本）
        ws.column_dimensions['A'].width = 19.0
        ws.column_dimensions['E'].width = 20.5
        ws.column_dimensions['F'].width = 19.0
        ws.column_dimensions['J'].width = 20.5

        # 頁面設定：A4 縱向、極小頁邊距（依照參考範本）
        ws.page_setup.paperSize = 9   # A4
        ws.page_setup.orientation = 'portrait'
        ws.page_margins.left   = 0.07874015748031496
        ws.page_margins.right  = 0.07874015748031496
        ws.page_margins.top    = 0.0
        ws.page_margins.bottom = 0.0

        label_font = Font(name='標楷體', size=14)
        addr_align = Alignment(horizontal='left',   vertical='bottom', wrap_text=True)
        name_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for pair_idx in range(0, len(data), 2):
            base_row = (pair_idx // 2) * 4 + 1
            left  = data[pair_idx]
            right = data[pair_idx + 1] if pair_idx + 1 < len(data) else None

            # 行高
            ws.row_dimensions[base_row    ].height = 39.75
            ws.row_dimensions[base_row + 1].height = 24.0
            ws.row_dimensions[base_row + 2].height = 33.75
            ws.row_dimensions[base_row + 3].height = 9.95

            # 合併儲存格
            ws.merge_cells(f'A{base_row}:E{base_row}')         # 左欄地址行
            ws.merge_cells(f'A{base_row + 1}:C{base_row + 1}') # 左欄姓名行
            ws.merge_cells(f'F{base_row}:J{base_row}')         # 右欄地址行
            ws.merge_cells(f'F{base_row + 1}:H{base_row + 1}') # 右欄姓名行

            for col_start, item in [('A', left), ('F', right)]:
                if item is None:
                    continue

                # 郵遞區號（查不到則省略）
                # 正規化：部分資料庫記錄使用 巿 (U+5DFF) 而非 市 (U+5E02)
                county_norm = item.get('county', '').replace('\u5dff', '\u5e02')
                town_norm   = item.get('town',   '').replace('\u5dff', '\u5e02')
                zip_code = _TAIWAN_POSTAL_CODES.get(county_norm, {}).get(town_norm, '')
                prefix = f"{zip_code}\u3000" if zip_code else ""

                # 地址（與 _build_applicant_address 相同邏輯：縣市+鄉鎮+村里+詳細地址）
                addr_parts = [item.get('county', ''), item.get('town', '')]
                if item.get('village'):
                    addr_parts.append(item['village'])
                addr_parts.append(item.get('address', ''))
                full_addr = prefix + "".join(addr_parts)

                # 姓名行：格式化案件編號 + 兩空格 + 申請人姓名
                name_line = (
                    f"{format_case_number(item.get('case_number', ''))}"
                    f"  {item.get('applicant_name', '')}"
                )

                # 寫入地址儲存格
                addr_cell = ws[f'{col_start}{base_row}']
                addr_cell.value     = full_addr
                addr_cell.font      = label_font
                addr_cell.alignment = addr_align

                # 寫入姓名儲存格
                name_cell = ws[f'{col_start}{base_row + 1}']
                name_cell.value     = name_line
                name_cell.font      = label_font
                name_cell.alignment = name_align

        output_path = self.temp_dir / f"address_labels_{year}.xlsx"
        wb.save(str(output_path))
        return str(output_path)

    # ─────────────────────────────────────────────────────────────────────────
    # 管路補助金額明細表（SDT）
    # ─────────────────────────────────────────────────────────────────────────

    def _write_sdt_sheet_header(self, ws, roc_year: int, office_name: str, start_row: int = 1) -> int:
        """
        寫入管路補助金額明細表的表頭（start_row 起 4 列），回傳第一筆資料列號。
        換頁時以目前的 cur_row 傳入。

        邊框策略：
        - 水平合併：先對每欄設 border 再 merge（同 _tpl_merge_horizontal 模式），
          確保合併範圍內各欄的 border XML 均已寫入
        - 垂直合併：起始格設 border + _tpl_force_interior_border 注入底格邊框
        - Row 1 標題列：A1/G1/H1 無框線；Q1/R1 以 _tpl_force_interior_border 注入上緣
        """
        cur_row = start_row

        # ── Row 1：標題列（A1/G1/H1 無框線）──────────────────────────────
        ws.row_dimensions[cur_row].height = 30.0

        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
        c = ws.cell(cur_row, 1)
        c.value     = '農業部農田水利署' + office_name
        c.font      = _SDT_TITLE_FONT
        c.alignment = _SDT_RT

        g1 = ws.cell(cur_row, 7)
        g1.value     = roc_year
        g1.font      = _SDT_TITLE_FONT
        g1.alignment = _SDT_LT

        ws.merge_cells(start_row=cur_row, start_column=8, end_row=cur_row, end_column=18)
        c = ws.cell(cur_row, 8)
        c.value     = '年度管路補助金額明細表'
        c.font      = _SDT_TITLE_FONT
        c.alignment = _SDT_LT

        # ── Row 2：空白列 ───────────────────────────────────────────────────
        r2 = cur_row + 1
        ws.row_dimensions[r2].height = 8.0

        # ── Row 3–4：欄位標題 ────────────────────────────────────────────
        r3 = cur_row + 2
        r4 = cur_row + 3
        ws.row_dimensions[r3].height = 20.0
        ws.row_dimensions[r4].height = 20.0

        def _hdr(row, col, text):
            c = ws.cell(row, col)
            c.value     = text
            c.font      = _SDT_HDR_FONT
            c.alignment = _SDT_CTR
            c.border    = _SDT_THIN_BORDER

        # 垂直合併（A3:A4 等）：起始格設 border，再 merge，再注入底格 border
        def _vmhdr(col, text):
            c = ws.cell(r3, col)
            c.value     = text
            c.font      = _SDT_HDR_FONT
            c.alignment = _SDT_CTR
            c.border    = _SDT_THIN_BORDER
            ws.merge_cells(start_row=r3, start_column=col, end_row=r4, end_column=col)
            self._tpl_force_interior_border(ws, r4, col, _SDT_THIN_BORDER)

        # 水平合併（G3:N3 等）：先對每欄設 border，再 merge
        def _hmhdr(col_start, col_end, text):
            for c in range(col_start, col_end + 1):
                ws.cell(r3, c).border = _SDT_THIN_BORDER
            ws.merge_cells(start_row=r3, start_column=col_start, end_row=r3, end_column=col_end)
            c = ws.cell(r3, col_start)
            c.value     = text
            c.font      = _SDT_HDR_FONT
            c.alignment = _SDT_CTR

        # Row 3 垂直合併欄（R3:R4）
        _vmhdr(1,  '設施編號')        # A3:A4
        _vmhdr(2,  '農戶姓名')        # B3:B4
        _vmhdr(3,  '面積\n(公頃)')           # C3:C4
        _vmhdr(4,  '地點')                # D3:D4
        _vmhdr(5,  '灌溉\n型式')            # D3:D4
        _vmhdr(6,  '農戶\n配合款')            # D3:D4
        _vmhdr(15, '工程費\n合計')    # O3:O4
        # 設計者欄（col 19）已隱藏，不輸出標頭

        # Row 3 水平合併欄
        _hmhdr(7,  14, '政府補助案')              # G3:N3
        _hmhdr(16, 18, '每公頃田間設施工程費價')  # P3:R3

        # Row 3 單欄標題
        # _hdr(r3, 3, '面積')    # C3
        # _hdr(r3, 5, '灌溉')    # E3
        # _hdr(r3, 6, '農戶')    # F3

        # Row 4 細項標題
        for col, text in (
            (7,  '末端設施'), (8, '水源設施'), (9, '調控設施'),
            (10, '蓄水池'), (11, '動力設備'), (12, '小計'),
            (13, '設計費'), (14, '總計'),
            (16, '補助費'), (17, '百分比'), (18, '總工程費'),
        ):
            _hdr(r4, col, text)

        return cur_row + 4  # 第一筆資料列

    def _write_sdt_data_row(self, ws, row_num: int, row_data: dict):
        """寫入單筆案件資料列（A–S，19 欄）"""
        ws.row_dimensions[row_num].height = 18.0

        values = [
            row_data.get('case_number', ''),
            row_data.get('applicant_name', ''),
            float(row_data.get('area_ha', 0) or 0),
            row_data.get('location', ''),
            row_data.get('irrigation_type', ''),
            row_data.get('farmer_contribution', 0),
            row_data.get('end_facility', 0),
            0,  # H 水源設施固定為 0
            row_data.get('control_facility', 0),
            row_data.get('reservoir', 0),
            row_data.get('power_equipment', 0),
            row_data.get('govt_subtotal', 0),
            row_data.get('design_fee', 0),
            row_data.get('total', 0),
            row_data.get('grand_total', 0),
            row_data.get('per_ha_subsidy', 0),
            row_data.get('per_ha_pct'),   # Q 百分比（P/R，Decimal ROUND_DOWN 至小數第二位）
            row_data.get('per_ha_grand_total', 0),
            # designer（col 19）已隱藏，不輸出
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row_num, col_idx)
            c.value  = val
            c.font   = _SDT_DATA_FONT
            c.border = _SDT_THIN_BORDER
            if col_idx == 3:
                c.number_format = _SDT_NUM_FMT_HA
                c.alignment = _SDT_RT
            elif 6 <= col_idx <= 16:
                c.number_format = _SDT_NUM_FMT_INT
                c.alignment = _SDT_RT
            elif col_idx == 17:
                c.number_format = _SDT_NUM_FMT_PCT
                c.alignment = _SDT_RT
            elif col_idx == 18:
                c.number_format = _SDT_NUM_FMT_R
                c.alignment = _SDT_RT
            elif col_idx in (1, 2, 4, 5):  # 設施編號、農戶姓名、地點、灌溉型式
                c.alignment = _SDT_CTR
            else:
                c.alignment = _SDT_LT

    @staticmethod
    def _build_irrigation_summaries(rows: list) -> dict:
        """
        依灌溉型式分組計算合計，回傳固定 5 個 key 的 dict：
        { '穿孔管': {...}, '噴頭': {...}, '滴灌': {...}, '微噴': {...}, '其它': {...} }
        """
        totals: dict = {t: {'count': 0, 'area_ha': 0.0,
                            'farmer_contribution': 0, 'end_facility': 0,
                            'water_source': 0, 'control_facility': 0,
                            'reservoir': 0, 'power_equipment': 0,
                            'govt_subtotal': 0, 'design_fee': 0,
                            'total': 0, 'grand_total': 0}
                        for t in _SDT_IRRIGATION_TYPES}
        for row in rows:
            irr = row.get('irrigation_type', '其它')
            if irr not in totals:
                irr = '其它'
            t = totals[irr]
            t['count']              += 1
            t['area_ha']            += float(row.get('area_ha', 0) or 0)
            t['farmer_contribution'] += int(row.get('farmer_contribution', 0))
            t['end_facility']        += int(row.get('end_facility', 0))
            t['water_source']        += 0
            t['control_facility']    += int(row.get('control_facility', 0))
            t['reservoir']           += int(row.get('reservoir', 0))
            t['power_equipment']     += int(row.get('power_equipment', 0))
            t['govt_subtotal']       += int(row.get('govt_subtotal', 0))
            t['design_fee']          += int(row.get('design_fee', 0))
            t['total']               += int(row.get('total', 0))
            t['grand_total']         += int(row.get('grand_total', 0))
        # 計算每公頃欄位（無條件捨去，取整數）與 Q 欄百分比
        for irr_type, t in totals.items():
            a = t['area_ha']
            if a > 0:
                _a = Decimal(str(a))
                t['per_ha_subsidy']     = int((Decimal(str(t['end_facility'])) / _a).to_integral_value(rounding=ROUND_DOWN))
                t['per_ha_grand_total'] = int((Decimal(str(t['grand_total']))   / _a).to_integral_value(rounding=ROUND_DOWN))
            else:
                t['per_ha_subsidy']     = 0
                t['per_ha_grand_total'] = 0
            # Q 欄：補助費 / 總工程費，精確至小數第二位，無條件捨去
            if t['per_ha_grand_total'] > 0:
                t['per_ha_pct'] = float(
                    (Decimal(str(t['per_ha_subsidy'])) / Decimal(str(t['per_ha_grand_total'])))
                    .quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                )
            else:
                t['per_ha_pct'] = None
        return totals

    def _write_sdt_summary_block(self, ws, start_row: int, summaries: dict, total_row: dict):
        """
        寫入固定 5 列灌溉型式合計 + 1 列總計（共 6 列）。
        start_row：合計區第一列（A欄顯示「合計」）。
        """
        ws.row_dimensions[start_row].height = 18.0
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row + 4, end_column=1
        )
        c = ws.cell(start_row, 1)
        c.value     = '合計'
        c.font      = _SDT_HDR_FONT
        c.alignment = _SDT_CTR
        c.border    = _SDT_THIN_BORDER

        for i, irr_type in enumerate(_SDT_IRRIGATION_TYPES):
            r = start_row + i
            ws.row_dimensions[r].height = 18.0
            t = summaries.get(irr_type, {})
            values = [
                None,  # A（合併）
                f"{t.get('count', 0)}件設施",
                t.get('area_ha', 0),
                None,  # D 地點空白
                irr_type,
                t.get('farmer_contribution', 0),
                t.get('end_facility', 0),
                0,
                t.get('control_facility', 0),
                t.get('reservoir', 0),
                t.get('power_equipment', 0),
                t.get('govt_subtotal', 0),
                t.get('design_fee', 0),
                t.get('total', 0),
                t.get('grand_total', 0),
                t.get('per_ha_subsidy', 0),
                t.get('per_ha_pct'),      # Q 百分比（Decimal ROUND_DOWN 至小數第二位）
                t.get('per_ha_grand_total', 0),
                # designer（col 19）已隱藏，不輸出
            ]
            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(r, col_idx)
                if col_idx == 1:
                    c.border = _SDT_THIN_BORDER
                    continue
                c.value  = val
                c.font   = _SDT_DATA_FONT
                c.border = _SDT_THIN_BORDER
                if col_idx == 3:
                    c.number_format = _SDT_NUM_FMT_HA
                    c.alignment = _SDT_RT
                elif 6 <= col_idx <= 16:
                    c.number_format = _SDT_NUM_FMT_INT
                    c.alignment = _SDT_RT
                elif col_idx == 17:
                    c.number_format = _SDT_NUM_FMT_PCT
                    c.alignment = _SDT_RT
                elif col_idx == 18:
                    c.number_format = _SDT_NUM_FMT_R
                    c.alignment = _SDT_RT
                else:
                    c.alignment = _SDT_CTR

        # 總計列
        total_r = start_row + 5
        ws.row_dimensions[total_r].height = 18.0
        ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=2)
        c = ws.cell(total_r, 1)
        c.value     = '總計'
        c.font      = _SDT_HDR_FONT
        c.alignment = _SDT_CTR
        c.border    = _SDT_THIN_BORDER

        total_values = {
            3:  total_row.get('area_ha', 0),
            6:  total_row.get('farmer_contribution', 0),
            7:  total_row.get('end_facility', 0),
            8:  0,
            9:  total_row.get('control_facility', 0),
            10: total_row.get('reservoir', 0),
            11: total_row.get('power_equipment', 0),
            12: total_row.get('govt_subtotal', 0),
            13: total_row.get('design_fee', 0),
            14: total_row.get('total', 0),
            15: total_row.get('grand_total', 0),
        }
        for col_idx in range(2, 19):  # col 19（設計者）已隱藏
            c = ws.cell(total_r, col_idx)
            c.border = _SDT_THIN_BORDER
            if col_idx in total_values:
                c.value = total_values[col_idx]
                c.font  = _SDT_DATA_FONT
                if col_idx == 3:
                    c.number_format = _SDT_NUM_FMT_HA
                    c.alignment = _SDT_RT
                else:
                    c.number_format = _SDT_NUM_FMT_INT
                    c.alignment = _SDT_RT

    @staticmethod
    def _set_sdt_footer(ws) -> None:
        """
        設定管路補助金額明細表的頁尾：
        - 全部職稱置於 left 區塊，以測試確認的固定空格數分隔
        - 前 8 個職稱間隔 26 個空格；主任工程師/副處長、副處長/處長 間隔 25 個空格
        - right 區塊：頁碼
        """
        sp26 = ' ' * 11
        sp25 = ' ' * 16
        titles = _SDT_SIGN_TITLES   # 9 個職稱
        body = (
            sp26.join(titles[:7])   # 灌推承辦人…主任工程師（前 7，間距 15×6）
            + sp25 + titles[7]      # 副處長（間距 20）
            + sp25 + titles[8]      # 處長（間距 20）
        )
        font = _SDT_FOOTER_FONT_TAG
        ws.oddFooter.left.text  = f'{font}{body}'
        ws.oddFooter.left.size  = 12
        ws.oddFooter.right.text = f'{font}   &P/&N'
        ws.oddFooter.right.size = 12

    async def generate_subsidy_details_list(
        self,
        grants_by_sheet: dict,
        year: str,
        office_name: str,
    ) -> str:
        """
        生成管路補助金額明細表 XLSX（3 個工作表）。

        Args:
            grants_by_sheet: {'農水署明細表': [...], '瑠公明細表': [...], '七星明細表': [...]}
            year: 民國年字串（如 '114'）
            office_name: 使用者所屬單位名稱（顯示於 A1:F1）

        Returns:
            str: 臨時檔案絕對路徑
        """
        roc_year = int(year)
        wb = Workbook()
        wb.remove(wb.active)  # 移除預設空白工作表

        for sheet_name in ['農水署明細表', '瑠公明細表', '七星明細表']:
            rows = grants_by_sheet.get(sheet_name, [])
            ws = wb.create_sheet(title=sheet_name)

            # 設定欄寬
            for col_idx, width in _SDT_COL_WIDTHS.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # 頁面設定（A4 橫向，明確縮放比）
            # 使用 scale 取代 fitToWidth=1：
            ws.page_setup.paperSize   = _SDT_PAPER_SIZE
            ws.page_setup.orientation = _SDT_ORIENTATION
            ws.page_setup.scale       = _SDT_SCALE

            # 邊距（單位：英寸）
            ws.page_margins.top    = 0.5
            ws.page_margins.bottom = 0.35  # 縮小下邊距，讓頁尾上移靠近內容
            ws.page_margins.left   = 0.5
            ws.page_margins.right  = 0.5
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.58   # 頁尾距頁底縮小，配合 bottom 調整

            # 開啟時使用整頁模式（Page Layout view）
            ws.sheet_view.view = 'pageLayout'
            # 關閉背景格線：資料列已有明確 border，不受影響；
            # 空白列因此不再顯示格線，防止「空白表格」在 Page Layout view 覆蓋頁尾
            ws.sheet_view.showGridLines = False

            # 分頁寫入資料列
            cur_row = self._write_sdt_sheet_header(ws, roc_year, office_name, start_row=1)
            data_rows_on_page = 0

            for row_data in rows:
                if data_rows_on_page >= _SDT_ROWS_PER_PAGE:
                    ws.row_breaks.append(Break(id=cur_row - 1))  # 手動分頁符
                    cur_row = self._write_sdt_sheet_header(ws, roc_year, office_name, start_row=cur_row)
                    data_rows_on_page = 0
                self._write_sdt_data_row(ws, cur_row, row_data)
                cur_row += 1
                data_rows_on_page += 1

            if rows:
                # 確保合計區整塊不跨頁
                if data_rows_on_page + _SDT_SUMMARY_ROWS > _SDT_ROWS_PER_PAGE:
                    ws.row_breaks.append(Break(id=cur_row - 1))
                    cur_row = self._write_sdt_sheet_header(ws, roc_year, office_name, start_row=cur_row)

                # 計算合計資料
                summaries = self._build_irrigation_summaries(rows)
                total_row = {
                    'area_ha':            sum(float(r.get('area_ha', 0) or 0) for r in rows),
                    'farmer_contribution': sum(int(r.get('farmer_contribution', 0)) for r in rows),
                    'end_facility':        sum(int(r.get('end_facility', 0)) for r in rows),
                    'control_facility':    sum(int(r.get('control_facility', 0)) for r in rows),
                    'reservoir':           sum(int(r.get('reservoir', 0)) for r in rows),
                    'power_equipment':     sum(int(r.get('power_equipment', 0)) for r in rows),
                    'govt_subtotal':       sum(int(r.get('govt_subtotal', 0)) for r in rows),
                    'design_fee':          sum(int(r.get('design_fee', 0)) for r in rows),
                    'total':               sum(int(r.get('total', 0)) for r in rows),
                    'grand_total':         sum(int(r.get('grand_total', 0)) for r in rows),
                }

                self._write_sdt_summary_block(ws, cur_row, summaries, total_row)
                last_row = cur_row + _SDT_SUMMARY_ROWS - 1
            else:
                last_row = cur_row - 1  # 無案件：僅標題列

            # 明確限定列印範圍，防止空白列在 Page Layout view 覆蓋頁尾
            ws.print_area = f'A1:R{last_row}'
            # 在最後一列後插入 row break，讓 Excel 明確知道這頁在此結束，
            # 避免 Page Layout view 將最後一頁延伸至預設頁面容量底部（多 ~7 列），
            # 覆蓋頁尾簽核欄位。
            ws.row_breaks.append(Break(id=last_row))

            self._set_sdt_footer(ws)

        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename  = f"subsidy_details_list_{year}_{timestamp}.xlsx"
        file_path = self.temp_dir / filename
        wb.save(str(file_path))
        return str(file_path)

    def cleanup_temp_files(self, max_age_hours: int = 24):
        """清理超過指定時間的臨時檔案"""
        import time

        current_time = time.time()
        max_age_seconds = max_age_hours * 3600

        for file_path in self.temp_dir.glob("*.xls*"):
            if current_time - file_path.stat().st_mtime > max_age_seconds:
                try:
                    file_path.unlink()
                except OSError:
                    pass  # 忽略刪除失敗的檔案